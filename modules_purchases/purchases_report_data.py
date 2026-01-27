# modules_purchases/purchases_report_data.py
from __future__ import annotations

import logging
import os
import json
import time
import datetime as dt
import asyncio
from typing import Dict, List, Any, Optional, Tuple
from copy import copy

import aiohttp
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from config_package import safe_read_json, safe_write_json

# Логирование
log = logging.getLogger("seller-bot.purchases_report_data")

# ─── Базовые пути / env ────────────────────────────────────────────────
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
load_dotenv(os.path.join(ROOT_DIR, ".env"))

# Единая схема путей (ТЗ 5.1)
try:
    from modules_common.paths import DATA_DIR, CACHE_PUR, ensure_dirs  # type: ignore
except Exception:
    DATA_DIR = os.path.join(ROOT_DIR, "data")
    CACHE_PUR = os.path.join(DATA_DIR, "cache", "purchases")
    os.makedirs(CACHE_PUR, exist_ok=True)

    def ensure_dirs():
        os.makedirs(DATA_DIR, exist_ok=True)
        os.makedirs(CACHE_PUR, exist_ok=True)


ensure_dirs()

OZON_CLIENT_ID = os.getenv("OZON_CLIENT_ID", "")
OZON_API_KEY = os.getenv("OZON_API_KEY", "")
OZON_COMPANY_ID = os.getenv("OZON_COMPANY_ID", "")

# WATCH_SKU используем для обязательного поля `skus` в /v1/analytics/stocks
WATCH_SKU: List[str] = [
    s.strip() for s in (os.getenv("WATCH_SKU", "") or "").split(",") if s.strip()
]

# ─── Конфигурация городов из .env ────────────────────────────────────────────


def _tokens_from_env(var: str, default_csv: str) -> Tuple[str, ...]:
    raw = os.getenv(var, default_csv) or default_csv
    toks = [t.strip().lower() for t in raw.replace(";", ",").split(",") if t.strip()]
    return tuple(dict.fromkeys(toks))  # уникальные, в исходном порядке


PURCHASES_CITY_COUNT = int(os.getenv("PURCHASES_CITY_COUNT", "2") or "2")
CITY1_NAME = (os.getenv("PURCHASES_CITY1_NAME", "Москва") or "Москва").strip()
CITY2_NAME = (os.getenv("PURCHASES_CITY2_NAME", "Хабаровск") or "Хабаровск").strip()
CITY1_TOKENS = _tokens_from_env("PURCHASES_CITY1_TOKENS", "моск,moscow,msk")
CITY2_TOKENS = _tokens_from_env("PURCHASES_CITY2_TOKENS", "хабар,khabarov,khv")


def _is_city1_title(t: str) -> bool:
    t = (t or "").strip().lower()
    return bool(t) and (t == CITY1_NAME.strip().lower() or any(tok in t for tok in CITY1_TOKENS))


def _is_city2_title(t: str) -> bool:
    """
    ВАЖНО: определяем второй город независимо от PURCHASES_CITY_COUNT.
    Это нужно, чтобы парсер корректно игнорировал «город 2» в конфигурации с 1 городом,
    когда пользователь прислал старый двухгородовой файл.
    """
    t = (t or "").strip().lower()
    return bool(t) and (t == CITY2_NAME.strip().lower() or any(tok in t for tok in CITY2_TOKENS))


# ─── Пути / кэш ──────────────────────────────────────────────────────────────
XLSX_FILENAME = os.getenv("PURCHASES_XLSX_NAME", "Товары.xlsx")
XLSX_PATH = os.path.join(DATA_DIR, XLSX_FILENAME)
PURCHASES_CACHE_FILE = os.path.join(CACHE_PUR, "purchases_cache.json")
STOCKS_CACHE_FILE = os.path.join(CACHE_PUR, "stocks_cache_purchases.json")
STOCKS_FALLBACK_JSON = os.path.join(DATA_DIR, "stocks_history.json")  # оффлайн fallback

# ⬇️ Путь и имя автогенерируемого ШАБЛОНА + «база» для копирования внешнего вида
TEMPLATE_FILENAME = os.getenv("PURCHASES_TEMPLATE_NAME", "Товары_шаблон.xlsx")
TEMPLATE_PATH = os.path.join(DATA_DIR, TEMPLATE_FILENAME)
# По умолчанию используем «Товары_шаблон.xlsx» как каркас (можно переопределить env-переменной)
TEMPLATE_BASE_NAME = os.getenv("PURCHASES_TEMPLATE_BASE", TEMPLATE_FILENAME)
TEMPLATE_BASE_PATH = (
    TEMPLATE_BASE_NAME
    if os.path.isabs(TEMPLATE_BASE_NAME)
    else os.path.join(DATA_DIR, TEMPLATE_BASE_NAME)
)

CACHE_TTL_HOURS = int(os.getenv("PURCHASES_CACHE_MAX_AGE_HOURS", "12"))
STOCKS_TTL_HOURS = int(os.getenv("STOCKS_CACHE_MAX_AGE_HOURS", "1"))

# API
OZON_STOCKS_URL = "https://api-seller.ozon.ru/v1/analytics/stocks"

# ── 6 метрик, как в «Отгрузках» ─────────────────────────────────────────────
STOCK_METRICS = [
    "checking",
    "in_transit",
    "valid_stock_count",
    "available_for_sale",
    "return_from_customer_stock_count",
    "reserved",
]

# ─── Вспомогательные утилиты ─────────────────────────────────────────────────


def _headers() -> Dict[str, str]:
    return {
        "Client-Id": OZON_CLIENT_ID,
        "Api-Key": OZON_API_KEY,
        "Content-Type": "application/json",
    }


def _is_fresh(saved_at_iso: str, ttl_hours: int) -> bool:
    try:
        saved = dt.datetime.fromisoformat(saved_at_iso)
        return (dt.datetime.now() - saved) <= dt.timedelta(hours=ttl_hours)
    except Exception:
        return False


def _read_cache(path: str) -> dict:
    return safe_read_json(path)


def _write_cache(path: str, payload: dict) -> None:
    safe_write_json(path, payload)


# ─── Нормализации ────────────────────────────────────────────────────────────


def _norm(s: Any) -> str:
    return str(s or "").strip().lower()


def _norm_status(s: Any) -> Optional[str]:
    """
    Унифицируем статусы к канону:
    Выкупаются → "Выкупаются"
    Доставляются → "Доставляются"
    Обрабатываются → "Обрабатываются"
    """
    raw = str(s or "").strip()
    if not raw:
        return None
    low = raw.lower()

    # точные канонические формы
    if low.startswith("выкуп"):
        return "Выкупаются"
    if low.startswith("достав"):
        return "Доставляются"
    if low.startswith("обраб"):
        return "Обрабатываются"

    # альтернативы
    if "в пути" in low:
        return "Доставляются"
    if "провер" in low:
        return "Обрабатываются"

    if raw in {"Выкупаются", "Доставляются", "Обрабатываются"}:
        return raw
    return None


def _to_int_safe(x: Any) -> int:
    """Терпеливое преобразование количеств с безопасной обработкой мусора/пустоты."""
    try:
        if isinstance(x, int):
            return max(0, int(x))
        if isinstance(x, float):
            return max(0, int(round(x)))
        s = str(x or "").strip().replace("\u00a0", " ").replace(" ", "")
        s = s.replace(",", ".")
        if s == "":
            return 0
        val = float(s)
        return max(0, int(round(val)))
    except Exception:
        return 0


def _sku_from_cell(val: Any) -> Optional[int]:
    """SKU читаем как целое число, поддерживаем строковые цифры и '1831342831.0'."""
    if val is None:
        return None
    if isinstance(val, int):
        return int(val)
    if isinstance(val, float):
        try:
            return int(round(val))
        except Exception:
            return None
    s = str(val).strip()
    if s.endswith(".0") and s.replace(".", "", 1).isdigit():
        try:
            return int(float(s))
        except Exception:
            pass
    if s.isdigit():
        return int(s)
    return None


# ─── Excel: распознавание заголовков ─────────────────────────────────────────
SKU_HEADERS = {"sku", "артикул", "product_id", "id", "артикул товара", "код"}
STATUS_HEADERS = {"статус", "статус заказа", "статус поставки", "состояние"}
QTY_HEADERS = {"кол-во", "количество", "qty", "quantity", "шт", "количество,шт", "количество, шт"}

# Детекторы по подстроке для "широкого" формата
STATUS_TOKENS = {
    "Выкупаются": ("выкуп",),
    "Доставляются": ("достав", "в пути"),
    "Обрабатываются": ("обраб", "провер"),
}

# Варианты «алиаса» (заголовок колонки с коротким именем)
# ⚠️ без «товар», чтобы не ловить статусные колонки
ALIAS_HEADERS = {"алиас", "alias", "наимен", "назван"}
NUMBER_TOKENS = ("номер", "№")


def _choose_sheet(wb) -> Worksheet:
    for name in wb.sheetnames:
        low = name.lower()
        if any(k in low for k in ("выкуп", "покуп", "товар", "sku", "product")):
            return wb[name]
    return wb.active


def _find_header_row(ws: Worksheet, max_scan_rows: int = 10) -> int:
    """
    Ищем строку заголовков.
    Устраивает наличие:
      - SKU и (Status и (Qty или города))
      ИЛИ
      - SKU и любая колонка, содержащая токены статусов (широкий вид).
    """
    for i in range(1, min(ws.max_row, max_scan_rows) + 1):
        titles = [_norm(c.value) for c in ws[i]]
        has_sku = any(t in SKU_HEADERS for t in titles)

        has_status = any(t in STATUS_HEADERS for t in titles)
        has_city1 = any(_is_city1_title(t) for t in titles)
        has_city2 = any(_is_city2_title(t) for t in titles)
        has_qty_like = any(t in QTY_HEADERS for t in titles) or has_city1 or has_city2

        has_wide_status = False
        for t in titles:
            if any(any(tok in t for tok in toks) for toks in STATUS_TOKENS.values()):
                has_wide_status = True
                break

        if has_sku and ((has_status and has_qty_like) or has_wide_status):
            return i
    return 1  # по умолчанию


def _build_header_maps(ws: Worksheet, header_row: int):
    """
    Возвращает:
      - base: dict ключевых колонок {'sku': j, 'status': j?, 'qty': j?, 'msk': j?, 'khv': j?, 'alias': j?, 'num': j?}
      - wide: словарь списков колонок по статусам {'Выкупаются':[j...], 'Доставляются':[...], 'Обрабатываются':[...]}
    """
    base: Dict[str, int] = {}
    wide = {"Выкупаются": [], "Доставляются": [], "Обрабатываются": []}

    for j, cell in enumerate(ws[header_row], start=1):
        t = _norm(cell.value)
        if not t:
            continue

        # базовые
        if t in SKU_HEADERS:
            base["sku"] = j
            continue
        if t in STATUS_HEADERS:
            base["status"] = j
            continue
        if t in QTY_HEADERS:
            base["qty"] = j
            continue

        # алиас/название
        if any(k in t for k in ALIAS_HEADERS):
            base["alias"] = j

        # порядковый номер (если есть)
        if any(tok in t for tok in NUMBER_TOKENS) or t == "№":
            base["num"] = j

        # города (как отдельные колонки)
        if _is_city1_title(t):
            base["msk"] = j
        if _is_city2_title(t):
            base["khv"] = j

        # широкий формат: любые колонки с токенами статусов
        for canon, toks in STATUS_TOKENS.items():
            if any(tok in t for tok in toks):
                wide[canon].append(j)

    return base, wide


def _qty_from_row_wide(ws: Worksheet, i: int, cols: List[int]) -> int:
    total = 0
    for j in cols:
        total += _to_int_safe(ws.cell(i, j).value)
    return total


def _qty_from_row_tall(ws: Worksheet, i: int, base: Dict[str, int]) -> int:
    if "qty" in base:
        return _to_int_safe(ws.cell(i, base["qty"]).value)
    # если явной qty нет — суммируем города
    total = 0
    if "msk" in base:
        total += _to_int_safe(ws.cell(i, base["msk"]).value)
    # второй город учитываем только при конфигурации на 2 города
    if PURCHASES_CITY_COUNT == 2 and "khv" in base:
        total += _to_int_safe(ws.cell(i, base["khv"]).value)
    return total


# ─── Основной парсер ─────────────────────────────────────────────────────────


def load_excel(force: bool = False) -> Dict[int, Dict[str, int]]:
    """
    Читает Excel и агрегирует по SKU и статусам.
    Поддерживает два вида таблиц:
      1) «Высокий»: SKU | Статус | Кол-во   (+ возможно {CITY1_NAME}/{CITY2_NAME} вместо Кол-во)
      2) «Широкий»: SKU | ... | [Выкупаются_*] | [Доставляются_*] | [Обрабатываются_*]
         где *_ — любые уточнения (в т.ч. имена/токены городов); все соответствующие колонки суммируются.
    Возврат: { sku: {"Выкупаются": n, "Доставляются": m, "Обрабатываются": k} }
    """
    file_sig = _xlsx_sig(XLSX_PATH)
    cached = _read_cache(PURCHASES_CACHE_FILE)
    if (not force) and _cache_is_valid(cached, CACHE_TTL_HOURS, file_sig):
        try:
            return {int(k): v for k, v in (cached.get("rows") or {}).items()}
        except Exception:
            pass

    if not os.path.exists(XLSX_PATH):
        _write_cache(
            PURCHASES_CACHE_FILE,
            {
                "saved_at": dt.datetime.now().isoformat(),
                "file_mtime": file_sig[0],
                "file_size": file_sig[1],
                "rows": {},
            },
        )
        return {}

    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = _choose_sheet(wb)
    header_row = _find_header_row(ws)
    base, wide = _build_header_maps(ws, header_row)

    # определяем режим: широкий (если есть хоть одна колонка статуса в wide), иначе высокий
    wide_mode = any(len(cols) > 0 for cols in wide.values())

    # Требуется хотя бы SKU
    if "sku" not in base:
        _write_cache(
            PURCHASES_CACHE_FILE,
            {
                "saved_at": dt.datetime.now().isoformat(),
                "file_mtime": file_sig[0],
                "file_size": file_sig[1],
                "rows": {},
            },
        )
        return {}

    agg: Dict[int, Dict[str, int]] = {}

    if wide_mode:
        # При конфигурации «один город» игнорируем колонки второго города, даже если они есть.
        if PURCHASES_CITY_COUNT == 1:
            filtered_wide: Dict[str, List[int]] = {}
            for st, cols in wide.items():
                keep: List[int] = []
                for j in cols:
                    title = str(ws.cell(header_row, j).value or "")
                    if not _is_city2_title(title):  # работает даже когда count=1
                        keep.append(j)
                filtered_wide[st] = keep or cols
            wide = filtered_wide

        # ── ШИРОКИЙ ФОРМАТ ──
        for i in range(header_row + 1, ws.max_row + 1):
            sku = _sku_from_cell(ws.cell(i, base["sku"]).value)
            if sku is None:
                continue

            buy_q = _qty_from_row_wide(ws, i, wide["Выкупаются"])
            del_q = _qty_from_row_wide(ws, i, wide["Доставляются"])
            proc_q = _qty_from_row_wide(ws, i, wide["Обрабатываются"])

            if buy_q + del_q + proc_q <= 0:
                continue

            slot = agg.setdefault(sku, {"Выкупаются": 0, "Доставляются": 0, "Обрабатываются": 0})
            slot["Выкупаются"] += buy_q
            slot["Доставляются"] += del_q
            slot["Обрабатываются"] += proc_q
    else:
        # ── ВЫСОКИЙ ФОРМАТ ── (нужен статус и qty/города)
        has_qty_like = (
            ("qty" in base) or ("msk" in base) or ("khv" in base and PURCHASES_CITY_COUNT == 2)
        )
        if "status" not in base or not has_qty_like:
            _write_cache(
                PURCHASES_CACHE_FILE,
                {
                    "saved_at": dt.datetime.now().isoformat(),
                    "file_mtime": file_sig[0],
                    "file_size": file_sig[1],
                    "rows": {},
                },
            )
            return {}

        for i in range(header_row + 1, ws.max_row + 1):
            sku = _sku_from_cell(ws.cell(i, base["sku"]).value)
            if sku is None:
                continue

            status_canon = _norm_status(ws.cell(i, base["status"]).value)
            if status_canon is None:
                continue

            qty = _qty_from_row_tall(ws, i, base)
            if qty <= 0:
                continue

            slot = agg.setdefault(sku, {"Выкупаются": 0, "Доставляются": 0, "Обрабатываются": 0})
            slot[status_canon] += qty

    _write_cache(
        PURCHASES_CACHE_FILE,
        {
            "saved_at": dt.datetime.now().isoformat(),
            "file_mtime": file_sig[0],
            "file_size": file_sig[1],
            "rows": agg,
        },
    )
    return agg


# ─── Вспомогательные для кэша/файла ──────────────────────────────────────────


def _xlsx_sig(path: str) -> Tuple[float, int]:
    try:
        st = os.stat(path)
        return st.st_mtime, st.st_size
    except Exception:
        return (0.0, 0)


def _cache_is_valid(cached: dict, ttl_hours: int, file_sig: Tuple[float, int]) -> bool:
    if not cached:
        return False
    if not _is_fresh(cached.get("saved_at", ""), ttl_hours):
        return False
    try:
        return float(cached.get("file_mtime", 0)) == float(file_sig[0]) and int(
            cached.get("file_size", 0)
        ) == int(file_sig[1])
    except Exception:
        return False


# ─── Stocks: формирование запроса и загрузка ─────────────────────────────────


def _payload_stocks(skus: List[int]) -> Dict[str, Any]:
    p: Dict[str, Any] = {
        "metrics": STOCK_METRICS,
        "dimension": ["sku"],
        "limit": 1000,
        "skus": [str(x) for x in skus[:100]],  # API ждёт строки; максимум 100 за запрос
    }
    if OZON_COMPANY_ID:
        p["company_id"] = OZON_COMPANY_ID
    return p


def _derive_skus_for_request(excel_rows: Dict[int, Dict[str, int]]) -> List[int]:
    if WATCH_SKU:
        only_digits = []
        for s in WATCH_SKU:
            s = s.strip()
            if s.isdigit():
                try:
                    only_digits.append(int(s))
                except Exception:
                    pass
        if only_digits:
            return only_digits
    if excel_rows:
        return list(excel_rows.keys())
    return []


def _load_stocks_from_fallback() -> Dict[int, Dict[str, float]]:
    js = safe_read_json(STOCKS_FALLBACK_JSON)
    if not js:
        return {}

    rows = js.get("rows")
    out: Dict[int, Dict[str, float]] = {}

    def _zero_six() -> Dict[str, float]:
        return {
            "available_for_sale": 0.0,
            "checking": 0.0,
            "in_transit": 0.0,
            "valid_stock_count": 0.0,
            "return_from_customer_stock_count": 0.0,
            "reserved": 0.0,
        }

    if isinstance(rows, dict):
        for k, v in rows.items():
            try:
                sku = int(k)
            except Exception:
                continue
            cur = _zero_six()
            cur["available_for_sale"] = float((v or {}).get("available_for_sale", 0) or 0)
            cur["checking"] = float((v or {}).get("checking", 0) or 0)
            cur["in_transit"] = float((v or {}).get("in_transit", 0) or 0)
            cur["valid_stock_count"] = float((v or {}).get("valid_stock_count", 0) or 0)
            cur["return_from_customer_stock_count"] = float(
                (v or {}).get("return_from_customer_stock_count", 0) or 0
            )
            cur["reserved"] = float((v or {}).get("reserved", 0) or 0)
            out[sku] = cur
        return out

    if isinstance(rows, list):
        for r in rows:
            try:
                sku = int(r.get("sku"))
            except Exception:
                continue
            cur = _zero_sекs()
            cur["available_for_sale"] = float(r.get("available_for_sale", 0) or 0)
            cur["checking"] = float(r.get("checking", 0) or 0)
            cur["in_transit"] = float(r.get("in_transit", 0) or 0)
            cur["valid_stock_count"] = float(r.get("valid_stock_count", 0) or 0)
            cur["return_from_customer_stock_count"] = float(
                r.get("return_from_customer_stock_count", 0) or 0
            )
            cur["reserved"] = float(r.get("reserved", 0) or 0)
            out[sku] = cur
        return out

    return {}


def _extract_stock_row(row: dict) -> Tuple[Optional[int], float, float, float, float, float, float]:
    """
    Возвращает кортеж:
      (sku,           available_for_sale,           checking,           in_transit,           valid_stock_count,           return_from_customer_stock_count,           reserved)
    Поддерживает top-level и metrics(list/dict) + альтернативные имена ключей.
    """
    sku_raw = row.get("sku") or row.get("product_id")
    try:
        sku = int(sku_raw)
    except Exception:
        sku = None

    # top-level
    a_top = row.get("available_stock_count")
    c_top = row.get("other_stock_count")
    t_top = row.get("transit_stock_count")
    v_top = row.get("valid_stock_count") or row.get("valid_stock") or row.get("valid")
    r_top = (
        row.get("return_from_customer_stock_count")
        or row.get("return_from_customer_stock")
        or row.get("return_from_customer")
    )
    rs_top = row.get("reserved") or row.get("reserved_stock")

    # metrics блок
    m = row.get("metrics") or row.get("value") or {}
    a = c = t = v = r = rs = None

    if isinstance(m, list):
        # предполагаем порядок как в STOCK_METRICS
        try:
            c = float(m[0]) if len(m) > 0 else None
        except Exception:
            pass
        try:
            t = float(m[1]) if len(m) > 1 else None
        except Exception:
            pass
        try:
            v = float(m[2]) if len(m) > 2 else None
        except Exception:
            pass
        try:
            a = float(m[3]) if len(m) > 3 else None
        except Exception:
            pass
        try:
            r = float(m[4]) if len(m) > 4 else None
        except Exception:
            pass
        try:
            rs = float(m[5]) if len(m) > 5 else None
        except Exception:
            pass
    elif isinstance(m, dict):

        def _fget(*keys):
            for k in keys:
                if k in m:
                    try:
                        return float(m.get(k) or 0)
                    except Exception:
                        return 0.0
            return None

        a = _fget("available_for_sale")
        c = _fget("checking")
        t = _fget("in_transit")
        v = _fget("valid_stock_count", "valid_stock", "valid")
        r = _fget(
            "return_from_customer_stock_count", "return_from_customer_stock", "return_from_customer"
        )
        rs = _fget("reserved", "reserved_stock")

    # финальные значения с fallback на top-level
    def _pick(x, x_top):
        try:
            return float(x if x is not None else (x_top or 0.0))
        except Exception:
            return 0.0

    a = _pick(a, a_top)
    c = _pick(c, c_top)
    t = _pick(t, t_top)
    v = _pick(v, v_top)
    r = _pick(r, r_top)
    rs = _pick(rs, rs_top)

    return sku, a, c, t, v, r, rs


async def load_stocks(force: bool = False) -> Dict[int, Dict[str, float]]:
    """Возвращает словарь: SKU -> { six metrics }."""
    cached = _read_cache(STOCKS_CACHE_FILE)
    if cached and not force and _is_fresh(cached.get("saved_at", ""), STOCKS_TTL_HOURS):
        try:
            out: Dict[int, Dict[str, float]] = {}
            for k, v in (cached.get("rows") or {}).items():
                sku = int(k)
                row = {
                    "available_for_sale": float((v or {}).get("available_for_sale", 0) or 0),
                    "checking": float((v or {}).get("checking", 0) or 0),
                    "in_transit": float((v or {}).get("in_transit", 0) or 0),
                    "valid_stock_count": float((v or {}).get("valid_stock_count", 0) or 0),
                    "return_from_customer_stock_count": float(
                        (v or {}).get("return_from_customer_stock_count", 0) or 0
                    ),
                    "reserved": float((v or {}).get("reserved", 0) or 0),
                }
                out[sku] = row
            return out
        except Exception:
            pass

    excel_rows = _read_cache(PURCHASES_CACHE_FILE).get("rows", {})
    try:
        excel_rows = {int(k): v for k, v in (excel_rows or {}).items()}
    except Exception:
        excel_rows = {}

    skus = _derive_skus_for_request(excel_rows)
    out: Dict[int, Dict[str, float]] = {}

    if skus:
        timeout = aiohttp.ClientTimeout(connect=5, total=30)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            for i in range(0, len(skus), 100):
                chunk = skus[i : i + 100]
                try:
                    body = _payload_stocks(chunk)
                    async with session.post(OZON_STOCKS_URL, headers=_headers(), json=body) as r:
                        if r.status == 429:
                            log.warning(f"API rate limit hit for stocks request (chunk {
                                    i // 100 + 1})")
                            await asyncio.sleep(2)  # Пауза при rate limit
                            continue

                        r.raise_for_status()
                        js = await r.json()
                        items = (
                            js.get("items")
                            or (js.get("result") or {}).get("data")
                            or js.get("data")
                            or []
                        )
                        log.debug(f"Successfully fetched stocks for {len(chunk)} SKUs")
                except asyncio.TimeoutError as e:
                    log.warning(f"Timeout fetching stocks (chunk {i // 100 + 1}): {e}")
                    continue
                except aiohttp.ClientError as e:
                    log.warning(f"Connection error fetching stocks (chunk {i // 100 + 1}): {e}")
                    continue
                except Exception as e:
                    log.error(f"Request error fetching stocks (chunk {
                            i // 100 + 1}): {e}", exc_info=True)
                    continue

                for row in items:
                    sku, a, c, t, v, rfc, rs = _extract_stock_row(row)
                    if sku is None:
                        continue
                    cur = out.setdefault(
                        sku,
                        {
                            "available_for_sale": 0.0,
                            "checking": 0.0,
                            "in_transit": 0.0,
                            "valid_stock_count": 0.0,
                            "return_from_customer_stock_count": 0.0,
                            "reserved": 0.0,
                        },
                    )
                    cur["available_for_sale"] += a
                    cur["checking"] += c
                    cur["in_transit"] += t
                    cur["valid_stock_count"] += v
                    cur["return_from_customer_stock_count"] += rfc
                    cur["reserved"] += rs

        if out:
            _write_cache(
                STOCKS_CACHE_FILE, {"saved_at": dt.datetime.now().isoformat(), "rows": out}
            )
            return out

    # fallback
    if isinstance(cached, dict) and cached.get("rows"):
        try:
            restored: Dict[int, Dict[str, float]] = {}
            for k, v in (cached.get("rows") or {}).items():
                sku = int(k)
                row = {
                    "available_for_sale": float((v or {}).get("available_for_sale", 0) or 0),
                    "checking": float((v or {}).get("checking", 0) or 0),
                    "in_transit": float((v or {}).get("in_transit", 0) or 0),
                    "valid_stock_count": float((v or {}).get("valid_stock_count", 0) or 0),
                    "return_from_customer_stock_count": float(
                        (v or {}).get("return_from_customer_stock_count", 0) or 0
                    ),
                    "reserved": float((v or {}).get("reserved", 0) or 0),
                }
                restored[sku] = row
            return restored
        except Exception:
            pass

    fb = _load_stocks_from_fallback()
    if fb:
        return fb

    return {}


# ============================================================================#
#                         ШАБЛОН «Товары_шаблон.xlsx»                         #
# ============================================================================#


def _env_order_list() -> List[str]:
    raw = os.getenv("WATCH_OFFERS", "") or ""
    return [s.strip() for s in raw.split(",") if s.strip()]


def _alias_map_from_env() -> Dict[int, str]:
    out: Dict[int, str] = {}
    for k, v in os.environ.items():
        if not k.startswith("ALIAS_"):
            continue
        try:
            sku = int(k.replace("ALIAS_", "", 1))
        except Exception:
            continue
        alias = str(v or "").strip()
        if alias:
            out[sku] = alias
    return out


def _watch_sku_list_int() -> List[int]:
    skus: List[int] = []
    for s in (os.getenv("WATCH_SKU", "") or "").split(","):
        s = (s or "").strip()
        if s.isdigit():
            try:
                skus.append(int(s))
            except Exception:
                continue
    # ещё ключи из ALIAS_*
    for k in os.environ.keys():
        if k.startswith("ALIAS_"):
            suf = k.replace("ALIAS_", "", 1)
            if suf.isdigit():
                try:
                    skus.append(int(suf))
                except Exception:
                    pass
    # уникально в исходном порядке
    seen, uniq = set(), []
    for x in skus:
        if x not in seen:
            seen.add(x)
            uniq.append(x)
    return uniq


def _order_key_for_row(sku: int, alias: str, env_order: List[str]) -> Tuple[int, int, str]:
    al = (alias or str(sku)).strip()
    try:
        idx = env_order.index(al)
        return (0, idx, al.lower())
    except ValueError:
        return (1, 10_000_000, al.lower())


def _status_title_for_city(status: str, city_name: str) -> str:
    if status == "Выкупаются":
        return f"Товар выкупается для ({city_name})"
    if status == "Доставляются":
        return f"Товар в доставке до ({city_name})"
    # "Обрабатываются"
    return f"Товар в обработке ФФ ({city_name})"


def _copy_column_appearance(ws: Worksheet, src_col: int, dst_col: int, header_row: int) -> None:
    """Копируем ширину столбца и стиль ячейки заголовка."""
    try:
        src_letter = get_column_letter(src_col)
        dst_letter = get_column_letter(dst_col)
        src_dim = ws.column_dimensions.get(src_letter)
        if src_dim and src_dim.width is not None:
            ws.column_dimensions[dst_letter].width = src_dim.width
    except Exception:
        pass
    try:
        src_cell = ws.cell(row=header_row, column=src_col)
        dst_cell = ws.cell(row=header_row, column=dst_col)
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
    except Exception:
        pass


def ensure_purchases_template(path: Optional[str] = None) -> str:
    """
    Генерирует/пересобирает Excel‑шаблон для загрузки выкупов.

    ✅ Сохраняем:
      • пустую 1-ю строку и пустую колонку слева (отступы),
      • названия и порядок столбцов,
      • ширины колонок и стили заголовков.

    Данные (строки) формируем из .env: WATCH_SKU + ALIAS_*.
    Порядок строк — как в WATCH_OFFERS. Колонка «Номер», если присутствует, нумеруется 1..N.

    Заголовки колонок приводим к унифицированному виду с подстановкой города:
      «… для (Город)», «… до (Город)», «… ФФ (Город)».

    ⚠️ Главное изменение:
      • Если ранее был режим «1 город» и колонки второго города были удалены,
        при переключении на «2 города» мы автоматически ДОБАВИМ их обратно.
    """
    ensure_dirs()
    tpl_path = path or TEMPLATE_PATH

    if not os.path.exists(TEMPLATE_BASE_PATH):
        raise FileNotFoundError(
            f"Не найден базовый файл для шаблона: {TEMPLATE_BASE_PATH}. "
            f"Положите исходный образец в data/ или задайте PURCHASES_TEMPLATE_BASE."
        )

    # Входные данные из окружения
    skus = _watch_sku_list_int()
    alias_map = _alias_map_from_env()
    env_order = _env_order_list()

    # Загружаем «базу» как каркас и приводим шапку к актуальной конфигурации городов
    wb = load_workbook(TEMPLATE_BASE_PATH)
    ws = _choose_sheet(wb)
    header_row = _find_header_row(ws)
    base, wide = _build_header_maps(ws, header_row)
    wide_mode = any(len(cols) > 0 for cols in wide.values())

    statuses = ("Выкупаются", "Доставляются", "Обрабатываются")

    if wide_mode:
        # ── ШИРОКИЙ ФОРМАТ ──
        # 1) При 1 городе — оставляем ровно по одной колонке на статус (Город1), остальные удаляем.
        # 2) При 2 городах — гарантируем наличие двух колонок (добавляем вторую при необходимости).
        for status in statuses:
            # Всегда пересобираем карту (индексы могли сдвигаться после операций)
            base, wide = _build_header_maps(ws, header_row)
            cols = list(sorted(wide.get(status, [])))
            if not cols:
                continue

            # Определяем текущие колонки по городам
            def _title(j: int) -> str:
                return str(ws.cell(row=header_row, column=j).value or "")

            city1_cols = [j for j in cols if _is_city1_title(_title(j))]
            city2_cols = [j for j in cols if _is_city2_title(_title(j))]
            c1 = city1_cols[0] if city1_cols else (cols[0] if cols else None)
            c2 = city2_cols[0] if city2_cols else None

            if PURCHASES_CITY_COUNT == 1:
                # Оставляем только город1
                if c1 is not None:
                    ws.cell(row=header_row, column=c1).value = _status_title_for_city(
                        status, CITY1_NAME
                    )
                # Удаляем избыточные одноимённые колонки
                extras = [j for j in cols if j != c1]
                for j in sorted(set(extras), reverse=True):
                    try:
                        ws.delete_cols(j)
                    except Exception:
                        pass
                # продолжаем к следующему статусу
                continue

            # PURCHASES_CITY_COUNT == 2
            # Гарантируем, что есть обе колонки; если второй нет — добавим её сразу за первой
            if c1 is None:
                # На всякий случай выберем первую доступную колонку как c1
                c1 = cols[0]
            if c2 is None:
                # Вставляем новую колонку сразу после c1
                insert_at = c1 + 1
                ws.insert_cols(insert_at)
                _copy_column_appearance(ws, c1, insert_at, header_row)
                c2 = insert_at
            # Переименуем и удалим лишние колонки, если их >2
            ws.cell(row=header_row, column=c1).value = _status_title_for_city(status, CITY1_NAME)
            ws.cell(row=header_row, column=c2).value = _status_title_for_city(status, CITY2_NAME)
            extras = [j for j in cols if j not in {c1, c2}]
            for j in sorted(set(extras), reverse=True):
                try:
                    ws.delete_cols(j)
                except Exception:
                    pass

        # После всех операций ещё раз пересоберём карту (для вставки строк данных ниже)
        base, wide = _build_header_maps(ws, header_row)

    else:
        # ── ВЫСОКИЙ ФОРМАТ ──
        if "msk" in base:
            ws.cell(row=header_row, column=base["msk"]).value = CITY1_NAME

        if PURCHASES_CITY_COUNT == 1:
            # При одном городе — удаляем столбец второго города (если был)
            if "khv" in base:
                ws.delete_cols(base["khv"])
                base, wide = _build_header_maps(ws, header_row)
        else:
            # Два города — гарантируем наличие второго городского столбца
            if "khv" not in base and "msk" in base:
                insert_at = base["msk"] + 1
                ws.insert_cols(insert_at)
                _copy_column_appearance(ws, base["msk"], insert_at, header_row)
                ws.cell(row=header_row, column=insert_at).value = CITY2_NAME
                base, wide = _build_header_maps(ws, header_row)
            elif "khv" in base:
                ws.cell(row=header_row, column=base["khv"]).value = CITY2_NAME

    # Очистим данные ниже шапки (перегенерируем)
    if ws.max_row > header_row:
        try:
            ws.delete_rows(header_row + 1, ws.max_row - header_row)
        except Exception:
            pass

    # Подготовим строки данных
    rows: List[Tuple[int, str]] = []
    for s in _watch_sku_list_int():
        rows.append((int(s), alias_map.get(int(s), str(s))))
    rows.sort(key=lambda t: _order_key_for_row(t[0], t[1], env_order))

    def _append_by_index(values: Dict[int, Any]) -> None:
        max_col = ws.max_column
        row_vals = [values.get(j, "") for j in range(1, max_col + 1)]
        ws.append(row_vals)

    has_num = "num" in base
    num_val = 1

    if wide_mode:
        # ── одна строка на SKU ──
        for sku, alias in rows:
            vals: Dict[int, Any] = {}
            if has_num and base.get("num"):
                vals[base["num"]] = num_val
            if "sku" in base:
                vals[base["sku"]] = sku
            if "alias" in base:
                vals[base["alias"]] = alias
            # инициализируем нулями все статусные колонки
            base, wide = _build_header_maps(ws, header_row)  # на всякий случай
            for j in wide.get("Выкупаются", []):
                vals[j] = 0
            for j in wide.get("Доставляются", []):
                vals[j] = 0
            for j in wide.get("Обрабатываются", []):
                vals[j] = 0
            _append_by_index(vals)
            num_val += 1
    else:
        # ── три строки на SKU ──
        for sku, alias in rows:
            for idx, st_name in enumerate(statuses):
                vals: Dict[int, Any] = {}
                if has_num and idx == 0 and base.get("num"):
                    vals[base["num"]] = num_val
                if "sku" in base:
                    vals[base["sku"]] = sku
                if "alias" in base:
                    vals[base["alias"]] = alias
                if "status" in base:
                    vals[base["status"]] = st_name
                if "qty" in base:
                    vals[base["qty"]] = 0
                if "msk" in base:
                    vals[base["msk"]] = 0
                if PURCHASES_CITY_COUNT == 2 and "khv" in base:
                    vals[base["khv"]] = 0
                _append_by_index(vals)
            num_val += 1

    os.makedirs(os.path.dirname(tpl_path), exist_ok=True)
    wb.save(tpl_path)
    return tpl_path
