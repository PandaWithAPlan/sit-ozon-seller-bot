# modules_sales/sales_buyout.py
"""
Отчёт по выкупам (подраздел «Продажи»):
- Принимаем 2 Excel-файла: «Заказы.xlsx» и «Отчет.xlsx»
- Из «Отчет.xlsx» собираем валидные трек-номера с учётом фильтров (исключаем возвраты,
    учитываем артикулы из .env WATCH_OFFERS, если указаны)
- В «Заказы.xlsx» подсвечиваем совпадающие трек-номера
- Возвращаем готовый файл с подсветкой и краткую статистику
"""

from __future__ import annotations

import os
import re
from typing import Optional, Tuple, Dict, Set

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ─────────────────────────────────────────────────────────────────────────────
# Базовые пути: пытаемся взять общий DATA_DIR, иначе — локальный fallback
# ─────────────────────────────────────────────────────────────────────────────
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
try:
    from modules_common.paths import DATA_DIR  # type: ignore
except Exception:
    DATA_DIR = os.path.join(ROOT_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)

UPLOAD_BASE_DIR = os.path.join(DATA_DIR, "uploads", "buyout")
os.makedirs(UPLOAD_BASE_DIR, exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# ENV: фильтр по артикулам из WATCH_OFFERS
# ─────────────────────────────────────────────────────────────────────────────
def _normalize_article_token(s: str) -> str:
    """Нормализуем артикул: тримминг, lower, убираем лишнее после ':' (если формат 'art:alias')."""
    s = (s or "").strip()
    if not s:
        return ""
    # допускаем формат "ARTICUL:alias" — берём левую часть
    s = s.split(":", 1)[0]
    return s.strip().lower()


def _env_watch_offers_set() -> Set[str]:
    raw = os.getenv("WATCH_OFFERS", "") or ""
    parts = re.split(r"[,\n;]+", raw)
    out: Set[str] = set()
    for p in parts:
        t = _normalize_article_token(p)
        if t:
            out.add(t)
    return out


WATCH_OFFERS_SET: Set[str] = _env_watch_offers_set()


# ─────────────────────────────────────────────────────────────────────────────
# Нормализация заголовков и утилиты
# ─────────────────────────────────────────────────────────────────────────────
def norm_header(s: str) -> str:
    """
    Унифицируем заголовки: убираем небуквенные символы, схлопываем пробелы, lower().
    """
    s = str(s or "")
    s = s.replace("\xa0", " ")
    s = re.sub(r"[^\w]+", " ", s, flags=re.U)  # всё, кроме букв/цифр/_
    s = re.sub(r"_+", " ", s)
    return s.strip().lower()


# ─────────────────────────────────────────────────────────────────────────────
# Классификация файла
# ─────────────────────────────────────────────────────────────────────────────
def classify_excel(path: str) -> Optional[str]:
    """
    Возвращает 'orders' или 'report' по структуре таблиц, иначе None.
    """
    try:
        sheets = pd.read_excel(path, sheet_name=None, nrows=300, engine="openpyxl")
    except Exception:
        return None

    # 1) признак "заказы": столбец "Номер отправления"
    for _, df in sheets.items():
        cols = [norm_header(c) for c in df.columns]
        if "номер отправления" in cols:
            return "orders"

    # 2) признак "отчет": где-то в первой колонке есть строка "№ п/п"
    for _, df in sheets.items():
        first_col = df.columns[0]
        series = df[first_col].astype(str).str.strip()
        if series.eq("№ п/п").any():
            return "report"

    return None


# ─────────────────────────────────────────────────────────────────────────────
# Парсинг «Отчёт.xlsx»
# ─────────────────────────────────────────────────────────────────────────────
def _build_clean_report_df(raw_df: pd.DataFrame) -> pd.DataFrame:
    """
    Нормализуем заголовки (многоуровневые) и обрезаем служебные строки.
    Ищем строку с '№ п/п' в первой колонке — это верх шапки.
    """
    header_row_idx = None
    for i in range(min(len(raw_df), 150)):
        if str(raw_df.iloc[i, 0]).strip() == "№ п/п":
            header_row_idx = i
            break
    if header_row_idx is None:
        raise Exception("Не найден заголовок '№ п/п' в отчёте.")

    # Первая и вторая строки шапки
    h1 = raw_df.iloc[header_row_idx]
    h2 = raw_df.iloc[header_row_idx + 1] if header_row_idx + 1 < len(raw_df) else None

    # Собираем названия колонок с «склейкой» уровня 1 и 2
    new_cols = []
    for col in raw_df.columns:
        t1 = h1[col]
        t2 = h2[col] if h2 is not None else None
        t1s = "" if pd.isna(t1) else str(t1).strip()
        t2s = "" if (t2 is None or pd.isna(t2)) else str(t2).strip()
        if t2s:
            if t1s and t1s != t2s:
                new_cols.append(f"{t1s}__{t2s}")  # например: "Отправление__Номер"
            else:
                new_cols.append(t2s)
        else:
            new_cols.append(t1s if t1s else str(col))

    # Данные обычно начинаются через 2–3 строки после первой шапки.
    start = header_row_idx + 2
    df = raw_df.iloc[start:].copy()
    df.columns = new_cols
    # если первая строка данных не числовая в "№ п/п" — смещаемся ещё на 1
    if "№ п/п" in df.columns:
        first_val = df.iloc[0]["№ п/п"]
        try:
            float(first_val)
        except Exception:
            start = header_row_idx + 3
            df = raw_df.iloc[start:].copy()
            df.columns = new_cols
    df.reset_index(drop=True, inplace=True)
    return df


def _detect_track_column(df: pd.DataFrame) -> str:
    """
    Находим колонку с трек-номерами по заголовкам (c нормализацией) либо по шаблону значений.
    """
    normalized_cols = {norm_header(c): c for c in df.columns}
    candidates = [
        "Отправление__Номер",
        "Отправление Номер",
        "Номер отправления",
        "Трек-номер",
        "Трек",
        "Tracking number",
        "Shipment id",
        "Номер отправки",
        "ID отправления",
    ]
    for cand in candidates:
        key = norm_header(cand)
        if key in normalized_cols:
            return normalized_cols[key]

    # Fallback по содержимому
    pattern = re.compile(r"^\d+(?:-\d+){1,3}$")
    best_col = None
    best_ratio = 0.0

    for col in df.columns:
        series = df[col].astype(str).str.strip()
        s_nonempty = series[series != ""]
        if s_nonempty.empty:
            continue
        ratio = s_nonempty.str.match(pattern).mean()  # доля «похоже на трек»
        if ratio > best_ratio:
            best_ratio = ratio
            best_col = col

    if best_col and best_ratio >= 0.15:
        return best_col

    raise Exception("Не найден столбец с трек-номером в 'Отчет.xlsx'.")


def parse_report(report_path: str, offers_filter_lower: Set[str]) -> Tuple[Set[str], Dict]:
    """
    Из отчёта собираем множество валидных треков + сводную статистику фильтров.
    Фильтр по артикулам берём из WATCH_OFFERS (offers_filter_lower).
    """
    sheets = pd.read_excel(report_path, sheet_name=None, engine="openpyxl")
    raw_df = next(iter(sheets.values()))  # первый лист
    df = _build_clean_report_df(raw_df)

    # строки с реальными данными (где «№ п/п» — число)
    if "№ п/п" in df.columns:
        mask_data = pd.to_numeric(df["№ п/п"], errors="coerce").notna()
    else:
        mask_data = pd.Series([True] * len(df))

    # исключаем возвраты
    if "Итого возвращено, руб." in df.columns:
        ret_series = pd.to_numeric(df["Итого возвращено, руб."], errors="coerce").fillna(0.0)
        mask_no_returns = ret_series <= 0
    else:
        mask_no_returns = pd.Series([True] * len(df))

    # фильтр по артикулам (если указан WATCH_OFFERS)
    if "Артикул" in df.columns and offers_filter_lower:
        art_series = df["Артикул"].astype(str).str.strip().str.lower()
        mask_art = art_series.isin(offers_filter_lower)
    else:
        mask_art = pd.Series([True] * len(df))

    mask_final = mask_data & mask_no_returns & mask_art

    stats = {
        "rows_total": int(mask_data.sum()),
        "excluded_returns": int((mask_data & ~mask_no_returns).sum()),
        "excluded_art": int((mask_data & mask_no_returns & ~mask_art).sum()),
    }

    # получаем колонку с трек-номером
    track_col = _detect_track_column(df)

    valid_tracks_series = (
        df.loc[mask_final, track_col].dropna().astype(str).map(lambda s: s.strip())
    )
    valid_tracks = {t for t in valid_tracks_series if t}

    stats["tracks_after_filters"] = len(valid_tracks)
    return valid_tracks, stats


# ─────────────────────────────────────────────────────────────────────────────
# Подсветка в «Заказы.xlsx»
# ─────────────────────────────────────────────────────────────────────────────
def highlight_matches_in_orders(
    orders_path: str, valid_tracks: Set[str], output_path: Optional[str] = None
) -> Tuple[str, int]:
    """
    Открывает «Заказы.xlsx», подсвечивает совпадения в колонке 'Номер отправления'.
    Возвращает (путь к файлу результата, количество совпадений).
    """
    wb = load_workbook(orders_path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

    # ищем индекс колонки с 'Номер отправления' (с запасом по синонимам)
    candidates = {
        "номер отправления",
        "трек",
        "трек номер",
        "tracking number",
        "shipment id",
    }
    col_index = None
    for idx, h in enumerate(headers):
        if norm_header(h) in candidates:
            col_index = idx
            break
    if col_index is None:
        # строгое название (на случай, если синонимы не сработали)
        try:
            col_index = headers.index("Номер отправления")
        except ValueError:
            raise Exception("Не найден столбец 'Номер отправления' в 'Заказы.xlsx'.")

    blue_fill = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")
    matched = 0

    for row in ws.iter_rows(min_row=2):
        cell = row[col_index]
        if cell.value:
            val = str(cell.value).strip()
            if val in valid_tracks:
                cell.fill = blue_fill
                matched += 1

    result_path = output_path or os.path.join(
        os.path.dirname(orders_path), "Результат_Заказы_C_Совпадениями.xlsx"
    )
    wb.save(result_path)
    return result_path, matched


# ─────────────────────────────────────────────────────────────────────────────
# Конвейер
# ─────────────────────────────────────────────────────────────────────────────
def process_files(
    orders_path: str, report_path: str, output_dir: Optional[str] = None
) -> Tuple[str, int, Dict]:
    """
    Основной конвейер: читаем отчёт (фильтруем), подсвечиваем в заказах.
    Фильтр по артикулам берём из WATCH_OFFERS.
    """
    valid_tracks, stats = parse_report(report_path, WATCH_OFFERS_SET)
    # результат складываем рядом с orders либо в переданную папку
    out_path = None
    if output_dir:
        out_path = os.path.join(output_dir, "Результат_Заказы_C_Совпадениями.xlsx")
    result_path, count = highlight_matches_in_orders(
        orders_path, valid_tracks, output_path=out_path
    )
    return result_path, count, stats


# ─────────────────────────────────────────────────────────────────────────────
# Инфо-текст
# ─────────────────────────────────────────────────────────────────────────────
def get_help_text(has_orders: bool = False, has_report: bool = False) -> str:
    return (
        "🧾 <b>Отчёт по выкупам</b>\n\n"
        "Пришлите <i>как документы</i> два файла Excel (.xlsx):\n"
        "  1) <b>Заказы.xlsx</b>\n"
        "  2) <b>Отчет.xlsx</b> (отчёт о реализации)\n\n"
        "Что делает отчёт:\n"
        "• Из «Отчёт.xlsx» берём трек‑номера, исключая возвраты;\n"
        "• Учитываем только артикулы из <code>WATCH_OFFERS</code> (.env), если заданы;\n"
        "• В «Заказы.xlsx» подсвечиваем совпадения 💙 и возвращаем файл‑результат."
    )
