# modules_shipments/shipments_need.py
from __future__ import annotations

import os
import math
import json
import re
import datetime as dt
from typing import Dict, Tuple, List, Optional, Any, DefaultDict
from collections import defaultdict

"""
Исправлено/доработано:
- `_demand_by_ws()` возвращает Dict[(wid, sku) -> D] и корректно тянет профили с view="warehouse".
- Во всех scope каждая строка теперь может содержать «sku_alias».
- Для scope='cluster' добавлены служебные поля: cid, dest="cluster", dest_name (имя кластера).
- Для scope='warehouse' добавлены служебные поля: wid, cid, dest="warehouse", dest_name (имя склада).
- Экспорт Excel добавляет явные колонки «SKU», «Товар», а во вкладках «Кластеры» и «Склады»
    также «Кластер»/«Склад» и «ID кластера»/«ID склада».
- На листах «Кластеры» и «Склады» реализована иерархическая свёртка: для каждого адреса итоговая
    строка + скрытые детальные строки SKU (раскрываются в Excel).
- v3.8: лист «Склады» переведён на 3‑уровневую свёртку «Кластер → Склад → SKU»,
    группы свёрнуты по умолчанию с «+», автофильтр, жирные строки‑итоги, сортировка
    (Отгрузить → Поддерживать → Распродать; по убыванию «Рекомендация, шт»).
"""

from .shipments_need_data import (
    # остатки
    get_stock6_by_sku,
    get_stock6_by_cluster,
    get_stock6_by_warehouse,
    # справочники / lead
    get_warehouses_map,
    get_lead_for_wid,
    # план 30
    load_plan30_by_sku,
    # настройки /warehouse
    get_demand_settings,
    # утилиты для шапки и закрытых складов
    get_wh_method_title,
    load_closed_warehouses,
)

# источник скоростей D — «Потребность»
try:
    from modules_shipments.shipments_demand import (  # type: ignore
        compute_D_average as _D_avg,
        compute_D_dynamics as _D_dyn,
        compute_D_hybrid as _D_hyb,
        compute_D_plan_distribution as _D_plan,
    )
except Exception:
    _D_avg = _D_dyn = _D_hyb = _D_plan = None  # type: ignore

# прямой фолбэк к нормализованным рядам
try:
    from modules_shipments.shipments_demand_data import (  # type: ignore
        fetch_sales_view as _dd_fetch_sales,
        sales_to_D_by_warehouse as _dd_sales_to_D_by_warehouse,
    )
except Exception:
    _dd_fetch_sales = None  # type: ignore

    def _dd_sales_to_D_by_warehouse(_rows, _period):  # type: ignore
        return {}


# алиасы SKU
try:
    from modules_sales.sales_facts_store import get_alias_for_sku  # type: ignore
except Exception:

    def get_alias_for_sku(sku: int) -> str:
        return str(sku)


# безопасный импорт метода плана (/method)
try:
    from modules_sales.sales_forecast import get_forecast_method_title as _plan_method_title  # type: ignore
except Exception:

    def _plan_method_title() -> str:
        return "/method: неизвестен"


# ⚠️ цели продаж (эффективный план)
try:
    from modules_sales.sales_goal import effective_plan30_by_sku as _effective_plan  # type: ignore
except Exception:

    def _effective_plan(plan: Dict[int, float], horizon_days: int = 30) -> Dict[int, float]:
        return dict(plan)


# ─────────────────────────────────────────────────────────────────────────────
# ENV / вспомогательные утилиты
# ─────────────────────────────────────────────────────────────────────────────
def _env_bool(val: str | None, default: bool = True) -> bool:
    if val is None or val == "":
        return default
    s = str(val).strip().lower()
    if s in {"1", "true", "yes", "y", "on", "да", "истина"}:
        return True
    if s in {"0", "false", "no", "n", "off", "нет"}:
        return False
    return default


ALERT_PLAN_HORIZON_DAYS = int(os.getenv("ALERT_PLAN_HORIZON_DAYS", "30"))
SHIP_ROUND_STEP = int(os.getenv("SHIP_ROUND_STEP", "2"))

# КОЭФФИЦИЕНТ СТРАХОВКИ (используем в цели Need от Плана)
SHIP_SAFETY_COEFF = float(os.getenv("SHIP_SAFETY_COEFF", "2.0"))

# Пороговая шкала обеспеченности R (по план-целям)
SHIP_RED_FACTOR_SHIP = float(os.getenv("SHIP_RED_FACTOR_SHIP", "1.5"))
SHIP_YELLOW_FACTOR_SHIP = float(os.getenv("SHIP_YELLOW_FACTOR_SHIP", "1.75"))
SHIP_GREEN_FACTOR_SHIP = float(os.getenv("SHIP_GREEN_FACTOR_SHIP", "2.0"))  # нижняя граница ENOUGH
SHIP_MAX_FACTOR_SHIP = float(
    os.getenv("SHIP_MAX_FACTOR_SHIP", "2.0")
)  # верхняя граница ENOUGH == 2.0

PROF_SHIP_GREEN_FACTOR_SHIP = float(os.getenv("PROF_SHIP_GREEN_FACTOR_SHIP", "2.0"))
PROF_SHIP_YELLOW_FACTOR_SHIP = float(os.getenv("PROF_SHIP_YELLOW_FACTOR_SHIP", "2.25"))
PROF_SHIP_RED_FACTOR_SHIP = float(os.getenv("PROF_SHIP_RED_FACTOR_SHIP", "2.5"))

# Фильтр «кривых» строк
EPS_STRICT = float(os.getenv("DEMAND_EPS_STRICT", "0.05"))

# Флаг светофорной индикации
SHIPMENTS_NEED_LIGHTS_ENABLED: bool = _env_bool(
    os.getenv("SHIPMENTS_NEED_LIGHTS_ENABLED"),
    default=True,
)

# Кэш плана
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
PLAN_CACHE_DIR = os.path.join(ROOT_DIR, "data", "cache", "shipments")
os.makedirs(PLAN_CACHE_DIR, exist_ok=True)
PLAN_CACHE_PATH = os.path.join(PLAN_CACHE_DIR, "plan30_cache.json")

# ── WATCH_SKU: порядок/множество «разрешённых» ──────────────────────────────


def _watch_sku_order_list() -> List[int]:
    raw = (os.getenv("WATCH_SKU", "") or "").replace("\n", ",").replace(" ", ",")
    out: List[int] = []
    seen: set[int] = set()
    for token in raw.split(","):
        token = token.strip()
        if not token:
            continue
        left = token.split(":", 1)[0].strip()
        try:
            s = int(left)
        except Exception:
            continue
        if s not in seen:
            seen.add(s)
            out.append(s)
    return out


def _watch_pos_map() -> Dict[int, int]:
    return {sku: i for i, sku in enumerate(_watch_sku_order_list())}


def _watch_set() -> set[int]:
    return set(_watch_sku_order_list())


WATCH_POS = _watch_pos_map()
WATCH_SET = _watch_set()

# ─────────────────────────────────────────────────────────────────────────────


def _round_to_step(x: float, step: int) -> int:
    if step <= 0:
        step = 1
    if x <= 0:
        return 0
    m = float(x) / float(step)
    k = int(math.floor(m + 0.5))
    if abs(m - (k - 0.5)) < 1e-12:
        k = int(math.ceil(m))
    out = int(k * step)
    if x > 0 and out == 0:
        out = step
    return out


def _sum6(m: Dict[str, float]) -> float:
    if not isinstance(m, dict):
        return 0.0
    s = 0.0
    for k in (
        "available_for_sale",
        "checking",
        "in_transit",
        "reserved",
        "return_from_customer_stock_count",
        "valid_stock_count",
    ):
        try:
            s += float(m.get(k, 0.0) or 0.0)
        except Exception:
            continue
    return s


def _as_int(x: Any) -> Optional[int]:
    try:
        return int(str(x).strip())
    except Exception:
        return None


def _safe_float(x: Any, default: float = 0.0) -> float:
    try:
        v = float(x)
        if math.isnan(v) or math.isinf(v):
            return default
        return v
    except Exception:
        return default


def _is_cruft_line(_title: str, _sku: int, d_day: float) -> bool:
    return float(d_day or 0.0) < EPS_STRICT


# ─────────────────────────────────────────────────────────────────────────────
# План 30 с кэш-фолбэком (+ цель продаж при необходимости)


def _read_json(path: str) -> dict:
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f) or {}
    except Exception:
        pass
    return {}


def _write_json(path: str, payload: dict) -> None:
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _apply_effective_plan(plan: Dict[int, float], *, horizon_days: int = 30) -> Dict[int, float]:
    try:
        return _effective_plan(plan, horizon_days=horizon_days)  # type: ignore
    except TypeError:
        try:
            return _effective_plan(plan)  # type: ignore
        except Exception:
            return dict(plan)
    except Exception:
        return dict(plan)


def _load_plan30_with_cache(*, use_goal: bool = False, horizon_days: int = 30) -> Dict[int, float]:
    current = load_plan30_by_sku() or {}
    cached = _read_json(PLAN_CACHE_PATH) or {}
    out: Dict[int, float] = {}
    for k in set(list(current.keys()) + list(cached.keys())):
        try:
            ki = int(k)
        except Exception:
            continue
        v = float(current.get(k, 0.0) or 0.0)
        if v <= 0.0:
            v = float(cached.get(str(ki), 0.0) or 0.0)
        if v > 0:
            out[ki] = v
    if out:
        _write_json(PLAN_CACHE_PATH, {str(k): float(v) for k, v in out.items()})
    return _apply_effective_plan(out, horizon_days=horizon_days) if use_goal else out


# ─────────────────────────────────────────────────────────────────────────────
# Светофор (R = Stock / NeedPlan)


def classify_status(R: float) -> Tuple[str, str, str]:
    R = _safe_float(R, 0.0)
    if R < SHIP_RED_FACTOR_SHIP:
        return ("DEFICIT", "strong", "🚚 Отгрузить")
    if R < SHIP_YELLOW_FACTOR_SHIP:
        return ("DEFICIT", "medium", "🚚 Отгрузить")
    if R < SHIP_GREEN_FACTOR_SHIP:
        return ("DEFICIT", "light", "🚚 Отгрузить")
    if math.isclose(R, SHIP_MAX_FACTOR_SHIP, rel_tol=1e-9, abs_tol=1e-9):
        return ("ENOUGH", "normal", "🔄 Поддерживать")
    if R < SHIP_MAX_FACTOR_SHIP:
        return ("DEFICIT", "light", "🚚 Отгрузить")
    if R <= PROF_SHIP_GREEN_FACTOR_SHIP:
        return ("SURPLUS", "light", "🏷 Распродать")
    if R <= PROF_SHIP_YELLOW_FACTOR_SHIP:
        return ("SURPLUS", "medium", "🏷 Распродать")
    return ("SURPLUS", "strong", "🏷 Распродать")


# ─────────────────────────────────────────────────────────────────────────────
# D / Lead / Stock


def _demand_by_ws(period_days: int, demand_method: str) -> Dict[Tuple[int, int], float]:
    """
    Возвращает карту D[(warehouse_id, sku)] -> D/день за указанный период.
    Сначала пытаемся использовать модуль modules_shipments.shipments_demand (payload с D_by_w_sku),
    затем — прямой фолбэк на нормализованные ряды.
    """
    # 1) Профиль потребности (payload)
    payload = None
    try:
        if demand_method == "plan_distribution" and _D_plan:
            payload = _D_plan(view="warehouse", period=period_days)  # type: ignore
        elif demand_method == "average" and _D_avg:
            payload = _D_avg(view="warehouse", period=period_days)  # type: ignore
        elif demand_method == "dynamics" and _D_dyn:
            payload = _D_dyn(view="warehouse", period=period_days)  # type: ignore
        elif _D_hyb:
            payload = _D_hyb(view="warehouse", period=period_days)  # type: ignore
    except TypeError:
        try:
            if demand_method == "plan_distribution" and _D_plan:
                payload = _D_plan(period=period_days)  # type: ignore
            elif demand_method == "average" and _D_avg:
                payload = _D_avg(period=period_days)  # type: ignore
            elif demand_method == "dynamics" and _D_dyn:
                payload = _D_dyn(period=period_days)  # type: ignore
            elif _D_hyb:
                payload = _D_hyb(period=period_days)  # type: ignore
        except Exception:
            payload = None

    if isinstance(payload, dict):
        d_map = payload.get("D_by_w_sku")
        if isinstance(d_map, dict) and d_map:
            return d_map

    # 2) Фолбэк: нормализованные строки → агрегирование до D[(wid, sku)]
    if _dd_fetch_sales:
        rows = _dd_fetch_sales(view="warehouse", days=period_days, force=False)
        d_map_fb = _dd_sales_to_D_by_warehouse(rows, period_days)
        if isinstance(d_map_fb, dict) and d_map_fb:
            return d_map_fb

    return {}


def _lead_for_wid_safe(wid: int) -> float:
    return float(get_lead_for_wid(int(wid)) or 0.0)


def _lead_weighted_for_sku(
    sku: int, D_ws: Dict[Tuple[int, int], float], subset_wids: Optional[List[int]] = None
) -> int:
    num = 0.0
    den = 0.0
    for (wid, s), d in D_ws.items():
        if s != sku:
            continue
        if subset_wids is not None and int(wid) not in subset_wids:
            continue
        L = _lead_for_wid_safe(int(wid))
        num += d * L
        den += d
    return int(round(num / den)) if den > 0 else 0


def _cid_name_map() -> Dict[int, str]:
    wm = get_warehouses_map()
    cid2name: Dict[int, str] = {}
    for _wid, (_wname, cid, cname) in wm.items():
        if int(cid or 0) and str(cname or "").strip() and int(cid) not in cid2name:
            cid2name[int(cid)] = str(cname).strip()
    return cid2name


# ─────────────────────────────────────────────────────────────────────────────
# Вспом. парсинг goal‑флагов


def _is_trueish(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return float(v) != 0.0
    s = str(v).strip().lower()
    return s in {"1", "true", "yes", "on", "goal", "targets"}


def _goal_mode_from_kwargs(kwargs: Dict[str, Any]) -> bool:
    if _is_trueish(kwargs.get("use_goal")):
        return True
    if _is_trueish(kwargs.get("use_sales_goal")):
        return True
    if _is_trueish(kwargs.get("goal")):
        return True
    if _is_trueish(kwargs.get("use_targets")):
        return True
    if _is_trueish(kwargs.get("use_sales_targets")):
        return True
    if str(kwargs.get("target", "")).strip().lower() == "goal":
        return True
    if str(kwargs.get("mode", "")).strip().lower() == "goal":
        return True
    return False


# ─────────────────────────────────────────────────────────────────────────────
# Карточка Plan-First


def _calc_need_line_planfirst(
    title: str,
    plan30: float,
    lead_days: int,
    stock: float,
    coef: float,
    D_day: float,
    S_days: int = 0,
) -> Dict[str, Any]:
    Plan30 = max(0.0, _safe_float(plan30))
    L = max(0, int(_safe_float(lead_days)))
    stock = max(0.0, _safe_float(stock))
    coef = max(0.0, _safe_float(coef, SHIP_SAFETY_COEFF))
    S = max(0, int(_safe_float(S_days, 0)))

    plan_day = Plan30 / 30.0
    H_eff = int(30) + L

    base_plan_units = plan_day * H_eff
    plan_q = int(round(base_plan_units))
    NeedPlan = max(0.0, float(plan_q) * coef)

    R = (stock / NeedPlan) if NeedPlan > 0 else 0.0
    block, sub, action = classify_status(R)

    if action == "🚚 Отгрузить":
        qty_raw = max(0.0, NeedPlan - stock)
    elif action == "🏷 Распродать":
        qty_raw = max(0.0, stock - NeedPlan)
    else:
        qty_raw = 0.0
    qty = _round_to_step(qty_raw, SHIP_ROUND_STEP)

    return {
        "title": title,
        "D": max(0.0, _safe_float(D_day)),
        "L": L,
        "S": S,
        "H_eff": H_eff,
        "Plan30": Plan30,
        "Coef": coef,
        "BaseNeed": float(plan_q),
        "UpperNeed": NeedPlan,
        "Stock": stock,
        "R": R,
        "block": block,
        "sub": sub,
        "action": action,
        "qty": int(qty),
        "qty_wh": 0,
        "qty_closed": 0,
        "qty_L": 0,
        "qty_S": 0,
        "ClosedWh": 0,
        "ClosedMark": 0,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Расчёт компонентов A→B→C→D на подмножестве складов


def _components_for_subset(
    sku: int,
    subset_wids: List[int],
    Plan30: float,
    Coef: float,
    S_days: int,
    D_ws: Dict[Tuple[int, int], float],
    stock6_wh: Dict[Tuple[int, str], Dict[int, Dict[str, float]]],
    closed_wids: set[int],
) -> Tuple[int, int, int, int]:
    stock_by_w: Dict[int, float] = {}
    D_by_w: Dict[int, float] = {}
    L_by_w: Dict[int, float] = {}

    for wid in subset_wids:
        Dw = sum(
            float(d or 0.0)
            for (wid2, s), d in D_ws.items()
            if int(wid2) == int(wid) and int(s) == int(sku)
        )
        if Dw <= 0:
            continue
        D_by_w[wid] = Dw
        L_by_w[wid] = _lead_for_wid_safe(int(wid))
        st = 0.0
        for (w_id, _wname), sku_map in (stock6_wh or {}).items():
            if int(w_id) != int(wid):
                continue
            for s2, m6 in (sku_map or {}).items():
                if _as_int(s2) == int(sku):
                    st += _sum6(m6)
        stock_by_w[wid] = st

    if not D_by_w:
        return (0, 0, 0, 0)

    D_total_subset = sum(D_by_w.values())
    D_total_sku = sum(float(d or 0.0) for (wid, s), d in D_ws.items() if int(s) == int(sku))
    share_subset = (D_total_subset / D_total_sku) if D_total_sku > 0 else 1.0

    plan30_q = int(round(_safe_float(Plan30)))
    Target_AB = float(plan30_q) * float(Coef) * share_subset

    stock_subset = sum(stock_by_w.values())
    Pool_AB = max(0.0, Target_AB - stock_subset)

    Gap_open = sum(
        max(0.0, D_by_w[wid] * 30.0 - stock_by_w.get(wid, 0.0))
        for wid in D_by_w
        if wid not in closed_wids
    )
    Gap_closed = sum(
        max(0.0, D_by_w[wid] * 30.0 - stock_by_w.get(wid, 0.0))
        for wid in D_by_w
        if wid in closed_wids
    )
    total_gap = Gap_open + Gap_closed

    qty_wh_A = qty_closed_A = 0.0
    if total_gap > 0 and Pool_AB > 0:
        take = min(Pool_AB, total_gap)
        share_open = (Gap_open / total_gap) if total_gap > 0 else 0.0
        qty_wh_A = take * share_open
        qty_closed_A = take - qty_wh_A

    rest_after_A = max(0.0, Pool_AB - (qty_wh_A + qty_closed_A))
    qty_wh_B = rest_after_A

    D30_sku = (float(Plan30) / 30.0) * share_subset
    D_open = sum(D_by_w[wid] for wid in D_by_w if wid not in closed_wids)
    D_closed = sum(D_by_w[wid] for wid in D_by_w if wid in closed_wids)
    if D_total_subset <= 0:
        open_share = closed_share = 0.0
    else:
        open_share = D_open / D_total_subset
        closed_share = D_closed / D_total_subset

    def _L_weighted(wids: List[int]) -> float:
        num = sum(D_by_w[w] * L_by_w[w] for w in wids if w in D_by_w)
        den = sum(D_by_w[w] for w in wids if w in D_by_w)
        return num / den if den > 0 else 0.0

    Lw_open = _L_weighted([w for w in D_by_w if w not in closed_wids])
    Lw_closed = _L_weighted([w for w in D_by_w if w in closed_wids])

    qty_L_total = D30_sku * (open_share * Lw_open + closed_share * Lw_closed)
    qty_S = D30_sku * open_share * float(max(0, int(S_days)))

    def R(x):
        return _round_to_step(max(0.0, x), SHIP_ROUND_STEP)

    return (R(qty_wh_A + qty_wh_B), R(qty_closed_A), R(qty_L_total), R(qty_S))


# ─────────────────────────────────────────────────────────────────────────────
# Главный расчёт


def compute_need(
    scope: str = "sku",
    dispatch_days: Optional[int] = None,
    *,
    filter_cluster: Optional[int] = None,
    filter_warehouse: Optional[int] = None,
    **goal_kwargs: Any,
) -> Dict[str, Any]:
    scope = str(scope or "sku").lower()
    now = dt.datetime.now().strftime("%d.%m.%Y %H:%М")

    demand_method, period_days = get_demand_settings(default_period=ALERT_PLAN_HORIZON_DAYS)
    D_ws = _demand_by_ws(period_days, demand_method)
    cid2name = _cid_name_map()
    wm = get_warehouses_map()

    stock6_sku = get_stock6_by_sku() or {}
    stock6_clu = get_stock6_by_cluster() or {}
    stock6_wh = get_stock6_by_warehouse() or {}

    goal_mode = _goal_mode_from_kwargs(goal_kwargs)
    plan30 = _load_plan30_with_cache(use_goal=goal_mode, horizon_days=30)
    closed_wids = set(load_closed_warehouses())

    filter_wids: Optional[List[int]] = None
    if filter_cluster is not None:
        filter_wids = [
            int(wid) for wid, (_wn, cid, _cn) in wm.items() if int(cid) == int(filter_cluster)
        ]
    elif filter_warehouse is not None:
        if int(filter_warehouse) in wm:
            filter_wids = [int(filter_warehouse)]
        else:
            filter_wids = []

    sku_set: set[int] = set(plan30.keys()) | set(int(k) for k in stock6_sku.keys())
    if WATCH_SET:
        sku_set &= WATCH_SET

    S_days = max(0, int(dispatch_days or 0))
    lines: List[Dict[str, Any]] = []

    for sku in sorted(sku_set):
        alias = (get_alias_for_sku(sku) or "").strip() or str(sku)

        wids_for_sku_all = [int(wid) for (wid, s) in D_ws.keys() if int(s) == int(sku)]
        if filter_wids is not None:
            wids_for_sku = [w for w in wids_for_sku_all if w in filter_wids]
        else:
            wids_for_sku = wids_for_sku_all[:]

        D_sku_all = sum(float(d or 0.0) for (wid, s), d in D_ws.items() if int(s) == int(sku))
        D_sku_subset = sum(float(D_ws.get((w, sku), 0.0)) for w in wids_for_sku)

        Plan30_full = float(plan30.get(sku, 0.0) or 0.0)
        Coef = float(SHIP_SAFETY_COEFF)

        if scope == "sku" and filter_wids is not None:
            if _is_cruft_line(alias, sku, D_sku_subset):
                continue
            share = (D_sku_subset / D_sku_all) if D_sku_all > 0 else 1.0
            Plan30_sub = Plan30_full * share
            L_eff = _lead_weighted_for_sku(sku, D_ws, subset_wids=wids_for_sku)
            stock_sub = 0.0
            for (w_id, _wname), sku_map in (stock6_wh or {}).items():
                if int(w_id) not in wids_for_sku:
                    continue
                for s2, m6 in (sku_map or {}).items():
                    if _as_int(s2) == int(sku):
                        stock_sub += _sum6(m6)

            base_line = _calc_need_line_planfirst(
                alias, Plan30_sub, L_eff, stock_sub, Coef, D_sku_subset, S_days
            )
            base_line["sku"] = int(sku)
            base_line["sku_alias"] = alias

            qwh, qcl, qL, qS = _components_for_subset(
                sku, wids_for_sku, Plan30_full, Coef, S_days, D_ws, stock6_wh, closed_wids
            )
            base_line["qty_wh"] = int(qwh)
            base_line["qty_closed"] = int(qcl)
            base_line["qty_L"] = int(qL)
            base_line["qty_S"] = int(qS)
            base_line["qty"] = int(qwh + qcl + qL + qS)
            closed_count = int(sum(1 for w in wids_for_sku if w in closed_wids))
            base_line["ClosedWh"] = int(closed_count)
            base_line["ClosedMark"] = int((closed_count > 0) or (qcl > 0))
            lines.append(base_line)
            continue

        if scope == "sku":
            if _is_cruft_line(alias, sku, D_sku_all):
                continue
            L_eff = _lead_weighted_for_sku(sku, D_ws)
            stock_total = _safe_float(
                _sum6(stock6_sku.get(sku, {}) if isinstance(stock6_sku.get(sku,                                                                               {}), dict) else {}),
                0.0,
            )
            base_line = _calc_need_line_planfirst(
                alias, Plan30_full, L_eff, stock_total, Coef, D_sku_all, S_days
            )
            base_line["sku"] = int(sku)
            base_line["sku_alias"] = alias

            wids_all = wids_for_sku_all
            qwh, qcl, qL, qS = _components_for_subset(
                sku, wids_all, Plan30_full, Coef, S_days, D_ws, stock6_wh, closed_wids
            )
            base_line["qty_wh"] = int(qwh)
            base_line["qty_closed"] = int(qcl)
            base_line["qty_L"] = int(qL)
            base_line["qty_S"] = int(qS)
            base_line["qty"] = int(qwh + qcl + qL + qS)
            closed_count = int(sum(1 for w in wids_all if w in closed_wids))
            base_line["ClosedWh"] = int(closed_count)
            base_line["ClosedMark"] = int((closed_count > 0) or (qcl > 0))
            lines.append(base_line)
            continue

        if scope == "cluster":
            cid_sum_D: DefaultDict[int, float] = defaultdict(float)
            for (wid, s), d in D_ws.items():
                if int(s) != int(sku):
                    continue
                cid = wm.get(int(wid), ("", 0, ""))[1]
                if cid:
                    cid_sum_D[int(cid)] += float(d or 0.0)

            for cid in sorted(cid_sum_D.keys()):
                cname = cid2name.get(int(cid), f"Кластер {cid}")
                if _is_cruft_line(cname, -1, cid_sum_D[cid]):
                    continue
                wids_in_c = [
                    int(wid)
                    for wid, (_wname, cid2, _cname) in wm.items()
                    if int(cid2) == int(cid)
                    and any(int(w) == int(wid) and int(s) == int(sku) for (w, s) in D_ws.keys())
                ]
                share_c = (cid_sum_D[cid] / D_sku_all) if D_sku_all > 0 else 0.0
                Plan30_c = Plan30_full * share_c

                if cid_sum_D[cid] > 0:
                    L_num = sum(
                        _lead_for_wid_safe(wid) * D_ws.get((wid, sku), 0.0) for wid in wids_in_c
                    )
                    L_den = sum(D_ws.get((wid, sku), 0.0) for wid in wids_in_c)
                    L_c = int(round(L_num / L_den)) if L_den > 0 else 0
                else:
                    L_c = 0

                stock_c = 0.0
                for (cid2, _cname), sku_map in (stock6_clu or {}).items():
                    if int(cid2) != int(cid):
                        continue
                    for s2, m6 in (sku_map or {}).items():
                        if _as_int(s2) == int(sku):
                            stock_c += _sum6(m6)

                line_c = _calc_need_line_planfirst(
                    title=cname,
                    plan30=Plan30_c,
                    lead_days=L_c,
                    stock=stock_c,
                    coef=Coef,
                    D_day=cid_sum_D[cid],
                    S_days=S_days,
                )
                line_c["sku"] = int(sku)
                line_c["sku_alias"] = alias
                line_c["cid"] = int(cid)
                line_c["dest"] = "cluster"
                line_c["dest_name"] = str(cname)

                qwh, qcl, qL, qS = _components_for_subset(
                    sku, wids_in_c, Plan30_full, Coef, S_days, D_ws, stock6_wh, closed_wids
                )
                line_c["qty_wh"] = int(qwh)
                line_c["qty_closed"] = int(qcl)
                line_c["qty_L"] = int(qL)
                line_c["qty_S"] = int(qS)
                line_c["qty"] = int(qwh + qcl + qL + qS)
                closed_count = int(sum(1 for w in wids_in_c if w in closed_wids))
                line_c["ClosedWh"] = int(closed_count)
                line_c["ClosedMark"] = int((closed_count > 0) or (qcl > 0))
                lines.append(line_c)
            continue

        if scope == "warehouse":
            D_w_sku: DefaultDict[int, float] = defaultdict(float)
            for (wid, s), d in D_ws.items():
                if int(s) != int(sku):
                    continue
                D_w_sku[int(wid)] += float(d or 0.0)

            D_total = sum(D_w_sku.values())
            for wid in sorted(D_w_sku.keys()):
                Dw = D_w_sku[wid]
                if Dw <= 0:
                    continue
                wname, cid, _cname = wm.get(int(wid), (f"Склад {wid}", 0, ""))
                stock_w = 0.0
                for (w_id, _wname), sku_map in (stock6_wh or {}).items():
                    if int(w_id) != int(wid):
                        continue
                    for s2, m6 in (sku_map or {}).items():
                        if _as_int(s2) == int(sku):
                            stock_w += _sum6(m6)
                Lw = int(_lead_for_wid_safe(wid))
                Plan30_w = (Plan30_full * (Dw / D_total)) if D_total > 0 else 0.0

                line_w = _calc_need_line_planfirst(
                    title=wname,
                    plan30=Plan30_w,
                    lead_days=Lw,
                    stock=stock_w,
                    coef=Coef,
                    D_day=Dw,
                    S_days=S_days,
                )
                line_w["sku"] = int(sku)
                line_w["sku_alias"] = alias
                line_w["wid"] = int(wid)
                line_w["cid"] = int(cid)
                line_w["dest"] = "warehouse"
                line_w["dest_name"] = str(wname)

                qwh, qcl, qL, qS = _components_for_subset(
                    sku, [wid], Plan30_full, Coef, S_days, D_ws, stock6_wh, closed_wids
                )
                line_w["qty_wh"] = int(qwh)
                line_w["qty_closed"] = int(qcl)
                line_w["qty_L"] = int(qL)
                line_w["qty_S"] = int(qS)
                line_w["qty"] = int(qwh + qcl + qL + qS)
                closed_count = 1 if wid in closed_wids else 0
                line_w["ClosedWh"] = int(closed_count)
                line_w["ClosedMark"] = int((closed_count > 0) or (qcl > 0))
                lines.append(line_w)

    groups = {"DEFICIT": [], "ENOUGH": [], "SURPLUS": []}

    def _bucket(ln: Dict[str, Any]) -> str:
        if ln.get("action") == "🔄 Поддерживать" and math.isclose(
            _safe_float(ln.get("R", 0.0)), SHIP_MAX_FACTOR_SHIP, rel_tol=1e-9, abs_tol=1e-9
        ):
            return "ENOUGH"
        if ln.get("sub") == "light":
            return "ENOUGH"
        return (
            "DEFICIT"
            if ln.get("action") == "🚚 Отгрузить"
            else ("SURPLUS" if ln.get("action") == "🏷 Распродать" else "ENOUGH")
        )

    for ln in lines:
        groups[_bucket(ln)].append(ln)

    def _env_idx(ln: Dict[str, Any]) -> int:
        try:
            return WATCH_POS.get(int(ln.get("sku")), 10**9)
        except Exception:
            return 10**9

    def _rank_deficit_color(ln: Dict[str, Any]) -> int:
        return {"strong": 0, "medium": 1}.get(str(ln.get("sub") or ""), 9)

    def _rank_surplus_color(ln: Dict[str, Any]) -> int:
        return {"strong": 0, "medium": 1}.get(str(ln.get("sub") or ""), 9)

    def _rank_enough_color(ln: Dict[str, Any]) -> int:
        blk = str(ln.get("block") or "")
        sub = str(ln.get("sub") or "")
        if blk == "DEFICIT" and sub == "light":
            return 0
        if blk == "ENOUGH":
            return 1
        if blk == "SURPLUS" and sub == "light":
            return 2
        return 9

    groups["DEFICIT"].sort(
        key=lambda ln: (_rank_deficit_color(ln), -int(ln.get("qty") or 0), _env_idx(ln))
    )
    groups["SURPLUS"].sort(
        key=lambda ln: (_rank_surplus_color(ln), -int(ln.get("qty") or 0), _env_idx(ln))
    )
    groups["ENOUGH"].sort(
        key=lambda ln: (_rank_enough_color(ln), -int(ln.get("qty") or 0), _env_idx(ln))
    )

    payload: Dict[str, Any] = {
        "updated_at": now,
        "scope": scope,
        "lines": lines,
        "groups": groups,
        "plan_method_title": _plan_method_title(),
        "wh_method": str(demand_method or "-"),
        "wh_period_days": int(period_days or ALERT_PLAN_HORIZON_DAYS),
        "dispatch_days": S_days,
        "goal_mode": int(1 if goal_mode else 0),
    }
    return payload


# ─────────────────────────────────────────────────────────────────────────────


def _strip_shipments_need_lights(text: str) -> str:
    try:
        marker = "Легенда:"
        idx = text.find(marker)
        if idx != -1:
            text = text[:idx].rstrip()
        for sign in ("🔴", "🟠", "🟢", "🟥", "🟨", "🟩", "✅"):
            text = text.replace(sign, "")
        text = re.sub(r"^[ \t]+", "", text, flags=re.MULTILINE)
        text = re.sub(r"\n{3,}", "\n\n", text)
        return text
    except Exception:
        return text


def format_need_text(payload: Dict[str, Any]) -> str:
    upd = payload.get("updated_at", "")
    groups = payload.get("groups") or {}
    plan_method_title = payload.get("plan_method_title") or _plan_method_title()
    wh_method_code = (payload.get("wh_method") or "-").strip().lower()
    wh_period_days = payload.get("wh_period_days") or ALERT_PLAN_HORIZON_DAYS
    goal_mode = bool(int(payload.get("goal_mode", 0)))

    wh_method_title = get_wh_method_title(wh_method_code)

    head = [
        "📊 ОТГРУЗКИ — РЕКОМЕНДАЦИИ",
        f"План продаж: {plan_method_title} (30 дн){' + цели' if goal_mode else ''}",
        f"Потребность складов (D): {wh_method_title} {wh_period_days} дн",
        f"⏱ Обновлено: {upd}",
        "",
    ]
    out: List[str] = []

    def _bullet(ln: Dict[str, Any]) -> str:
        if ln.get("block") == "DEFICIT":
            return (
                "🔴" if ln.get("sub") == "strong" else ("🟠" if ln.get("sub") == "medium" else "🟢")
            )
        if ln.get("block") == "SURPLUS":
            return (
                "🟥" if ln.get("sub") == "strong" else ("🟨" if ln.get("sub") == "medium" else "🟩")
            )
        return "✅"

    def _has_closed(ln: Dict[str, Any]) -> bool:
        try:
            return bool(
                int(
                    ln.get("ClosedMark")
                    or ln.get("ClosedWh")
                    or (1 if int(ln.get("qty_closed") or 0) > 0 else 0)
                )
            )
        except Exception:
            return False

    def _title(ln: Dict[str, Any]) -> str:
        return str(ln.get("title") or "")

    def _act_line(ln: Dict[str, Any]) -> str:
        upper_now = int(round(_safe_float(ln.get("UpperNeed", 0.0))))
        stock_now = int(round(_safe_float(ln.get("Stock", 0.0))))
        return f"🛒 У Ozon {stock_now} шт • " f"📊 Необходимо {upper_now} шт"

    def _detail(ln: Dict[str, Any]) -> str:
        qty_total = int(ln.get("qty") or 0)
        q_wh = int(ln.get("qty_wh") or 0)
        q_cl = int(ln.get("qty_closed") or 0)
        q_L = int(ln.get("qty_L") or 0)
        q_S = int(ln.get("qty_S") or 0)
        return f"🚛 Рекомендации: {qty_total} " f"(🚚 {q_wh}; 🚫 {q_cl}; ⏱ {q_L}; 🗓 {q_S})"

    def _row(ln: Dict[str, Any]) -> List[str]:
        lines_block: List[str] = []
        b = _bullet(ln)
        if b == "✅" and _has_closed(ln):
            b = "🚫"
        lines_block.append(f"{b} {_title(ln)}:")
        lines_block.append(_act_line(ln))
        lines_block.append(_detail(ln))
        lines_block.append("")
        return lines_block

    def _emit_block(title: str, rows: List[Dict[str, Any]]):
        part = [title, ""]
        if not rows:
            status_name = title.replace("🔻 ", "").replace("🔺 ", "").replace("✅ ", "")
            part.append(f"• Нет позиций в статусе «{status_name}».")
            part.append("")
            return part
        for ln in rows[:10000]:
            part += _row(ln)
        return part

    out += _emit_block("✅ Достаточно", groups.get("ENOUGH", []))
    out += _emit_block("🔻 Дефицит", groups.get("DEFICIT", []))
    out += _emit_block("🔺 Профицит", groups.get("SURPLUS", []))

    legend = [
        "",
        "Легенда:",
        "🚛 — суммарная рекомендация (итого)",
        "🚚 — на открытые склады (ямы + подушка)",
        "🚫 — за закрытые склады (перераспределение/участие закрытых складов)",
        "⏱ — покрытие доставки L (доп. объём сверх плана)",
        "🗓 — покрытие срока до отгрузки S (доп. объём сверх плана)",
        "✅ — норма (R=2.0, ничего не делать)",
        "🟥🟨🟩 — профицит (распродать до цели)",
        "🔴🟠🟢 — дефицит (отгрузить до цели)",
    ]

    full_text = "\n".join(head + out + legend)
    if not SHIPMENTS_NEED_LIGHTS_ENABLED:
        full_text = _strip_shipments_need_lights(full_text)
    return full_text


# ─────────────────────────────────────────────────────────────────────────────


def _normalize_view_to_scope(view: Optional[str], scope: Optional[str]) -> str:
    if scope:
        return str(scope).strip().lower()
    v = (view or "sku").strip().lower()
    if v in ("sku", "items", "товары"):
        return "sku"
    if v in ("cluster", "clusters", "кластер", "кластеры"):
        return "cluster"
    if v in ("warehouse", "warehouses", "склад", "склады"):
        return "warehouse"
    return "sku"


def need_to_ship_text(
    view: str = "sku",
    *,
    scope: Optional[str] = None,
    dispatch_days: Optional[int] = None,
    filter_cluster: Optional[int] = None,
    filter_warehouse: Optional[int] = None,
    **goal_kwargs: Any,
) -> str:
    try:
        eff_scope = _normalize_view_to_scope(view, scope)
        payload = compute_need(
            scope=eff_scope,
            dispatch_days=dispatch_days,
            filter_cluster=filter_cluster,
            filter_warehouse=filter_warehouse,
            **goal_kwargs,
        )
        return format_need_text(payload)
    except Exception:
        now = dt.datetime.now().strftime("%d.%m.%Y %H:%M")
        dummy = {
            "updated_at": now,
            "groups": {"DEFICIT": [], "ENOUGH": [], "SURPLUS": []},
            "plan_method_title": _plan_method_title(),
            "wh_method": "average",
            "wh_period_days": ALERT_PLAN_HORIZON_DAYS,
            "dispatch_days": int(dispatch_days or 0),
            "scope": "sku",
            "lines": [],
            "goal_mode": 0,
        }
        return format_need_text(dummy)


def shipments_text(**kwargs) -> str:
    return need_to_ship_text(**kwargs)


def need_text(**kwargs) -> str:
    return need_to_ship_text(**kwargs)


# ─────────────────────────────────────────────────────────────────────────────


def export_need_excel(path: str, data_sku: dict, data_cluster: dict, data_wh: dict) -> str:
    import pandas as pd  # type: ignore

    try:
        if os.path.exists(path):
            os.remove(path)
    except Exception:
        pass

    def _to_df(payload: dict) -> "pd.DataFrame":
        scope = str((payload.get("scope") or "sku")).strip().lower()
        rows = payload.get("lines") or []

        # Подготовим данные + «человеческие» алиасы для SKU
        norm_rows: List[dict] = []
        for r in rows:
            rr = dict(r)
            try:
                s = int(rr.get("sku"))
            except Exception:
                s = None
            alias = ""
            if s is not None:
                try:
                    alias = (get_alias_for_sku(s) or "").strip()
                except Exception:
                    alias = ""
            rr["sku_alias"] = alias or (str(s) if s is not None else "")
            norm_rows.append(rr)

        # Колонки зависят от scope (адрес назначения явно)
        if scope == "warehouse":
            cols = [
                ("SKU", "sku"),
                ("Товар", "sku_alias"),
                ("Склад", "title"),
                ("ID склада", "wid"),
                ("ID кластера", "cid"),
                ("Скорость, шт/день", "D"),
                ("Срок доставки, дн", "L"),
                ("Срок до отгрузки, дн", "S"),
                ("План (30 дн), шт", "Plan30"),
                ("Коэффициент", "Coef"),
                ("База 30 дней, шт", "BaseNeed"),
                ("Необходимо, шт", "UpperNeed"),
                ("Запас, шт", "Stock"),
                ("Обеспеченность, R", "R"),
                ("Действие", "action"),
                ("Рекомендация, шт", "qty"),
                ("На открытые, шт", "qty_wh"),
                ("За закрытые, шт", "qty_closed"),
                ("На лаг L, шт", "qty_L"),
                ("На лаг S, шт", "qty_S"),
                ("Закрытых складов, шт", "ClosedWh"),
            ]
        elif scope == "cluster":
            cols = [
                ("SKU", "sku"),
                ("Товар", "sku_alias"),
                ("Кластер", "title"),
                ("ID кластера", "cid"),
                ("Скорость, шт/день", "D"),
                ("Срок доставки, дн", "L"),
                ("Срок до отгрузки, дн", "S"),
                ("План (30 дн), шт", "Plan30"),
                ("Коэффициент", "Coef"),
                ("База 30 дней, шт", "BaseNeed"),
                ("Необходимо, шт", "UpperNeed"),
                ("Запас, шт", "Stock"),
                ("Обеспеченность, R", "R"),
                ("Действие", "action"),
                ("Рекомендация, шт", "qty"),
                ("На открытые, шт", "qty_wh"),
                ("За закрытые, шт", "qty_closed"),
                ("На лаг L, шт", "qty_L"),
                ("На лаг S, шт", "qty_S"),
                ("Закрытых складов, шт", "ClosedWh"),
            ]
        else:  # scope == "sku"
            cols = [
                ("SKU", "sku"),
                ("Товар", "sku_alias"),
                ("Скорость, шт/день", "D"),
                ("Срок доставки, дн", "L"),
                ("Срок до отгрузки, дн", "S"),
                ("План (30 дн), шт", "Plan30"),
                ("Коэффициент", "Coef"),
                ("База 30 дней, шт", "BaseNeed"),
                ("Необходимо, шт", "UpperNeed"),
                ("Запас, шт", "Stock"),
                ("Обеспеченность, R", "R"),
                ("Действие", "action"),
                ("Рекомендация, шт", "qty"),
                ("На открытые, шт", "qty_wh"),
                ("За закрытые, шт", "qty_closed"),
                ("На лаг L, шт", "qty_L"),
                ("На лаг S, шт", "qty_S"),
                ("Закрытых складов, шт", "ClosedWh"),
            ]

        def _clean_action(a: str) -> str:
            a = str(a or "")
            if "Отгрузить" in a:
                return "Отгрузить"
            if "Распродать" in a:
                return "Распродать"
            if "Поддерживать" in a:
                return "Поддерживать"
            return a

        data = []
        for r in norm_rows:
            row = {}
            for ru, key in cols:
                val = r.get(key)
                if key == "action":
                    val = _clean_action(val)
                row[ru] = val
            data.append(row)

        import pandas as pd  # type: ignore

        df = pd.DataFrame(data, columns=[ru for ru, _ in cols])
        # Маркер закрытого склада — только на листе «Склады»
        try:
            scope = str((payload.get("scope") or "sku")).strip().lower()
            if scope == "warehouse":

                def _is_closed(v) -> bool:
                    try:
                        return int(v or 0) > 0
                    except Exception:
                        return False

                df["Маркер закрытого склада"] = df["Закрытых складов, шт"].apply(
                    lambda v: "🚫" if _is_closed(v) else ""
                )
        except Exception:
            pass

        return df

    # --- сортировка внутри групп и свёртка (2-уровневая, для листа «Кластеры»)
    def _with_outline(
        df_src, group_col: str, sum_cols: List[str]
    ) -> Tuple["pd.DataFrame", List[int]]:
        import pandas as pd  # type: ignore

        if group_col not in df_src.columns:
            return df_src, [0] * len(df_src)

        # сортировка по действию и рекомендации
        order_action = {"Отгрузить": 0, "Поддерживать": 1, "Распродать": 2}
        if "Действие" in df_src.columns and "Рекомендация, шт" in df_src.columns:
            df_src = df_src.copy()
            df_src["__act"] = df_src["Действие"].map(order_action).fillna(9)
            df_src.sort_values(
                by=[group_col, "__act", "Рекомендация, шт"],
                ascending=[True, True, False],
                inplace=True,
            )
            df_src.drop(columns=["__act"], inplace=True, errors="ignore")

        out_rows: List[dict] = []
        levels: List[int] = []
        for gval, gdf in df_src.groupby(group_col, sort=True):
            # строка итога группы (уровень 0)
            header = {col: "" for col in df_src.columns}
            header[group_col] = gval
            if "Товар" in header:
                header["Товар"] = f"{gval} — итого"
            for c in sum_cols:
                if c in gdf.columns:
                    try:
                        header[c] = float(gdf[c].fillna(0).sum())
                    except Exception:
                        header[c] = None
            out_rows.append(header)
            levels.append(0)

            # детальные строки (уровень 1)
            for _, r in gdf.iterrows():
                out_rows.append({col: r.get(col) for col in df_src.columns})
                levels.append(1)

        df_out = pd.DataFrame(out_rows, columns=list(df_src.columns))
        return df_out, levels

    def _append_grand_total(df, sum_cols: List[str]):
        import pandas as pd  # type: ignore

        total = {col: "" for col in df.columns}
        label_col = "Товар" if "Товар" in df.columns else df.columns[0]
        total[label_col] = "ИТОГО ПО ЛИСТУ"
        for c in sum_cols:
            if c in df.columns:
                try:
                    total[c] = float(pd.to_numeric(df[c], errors="coerce").fillna(0).sum())
                except Exception:
                    total[c] = None
        return pd.concat([df, pd.DataFrame([total])], ignore_index=True)

    def _write_excel(engine: str) -> None:
        import pandas as pd  # type: ignore

        with pd.ExcelWriter(path, engine=engine) as wr:
            meta = {
                "Название отчёта": "ОТГРУЗКИ — РЕКОМЕНДАЦИИ",
                "Версия формы": "ТЗ v3.8 (Кластер→Склад→SKU,                                              свёртки по умолчанию,                                              автофильтр,                                              жирные итоги,                                              сортировка)",
                "План продаж": str(
                    (data_sku or {}).get("plan_method_title")
                    or (data_cluster or {}).get("plan_method_title")
                    or ""
                ),
                "Методика потребности": get_wh_method_title(
                    str((data_sku or {}).get("wh_method") or "")
                ),
                "Период потребности, дн": int((data_sku or {}).get("wh_period_days") or 30),
                "S (дни до отгрузки)": int((data_sku or {}).get("dispatch_days") or 0),
                "Дата формирования": str(
                    (data_sku or {}).get("updated_at")
                    or dt.datetime.now().strftime("%d.%m.%Y %H:%M")
                ),
            }
            pd.DataFrame([meta]).to_excel(wr, sheet_name="Сводка", index=False)

            # Детальные DF
            df_sku = _to_df(data_sku)
            df_cluster_detail = _to_df(data_cluster)
            df_wh_detail = _to_df(data_wh)

            # ➕ Добавляем человекочитаемый «Кластер» в df_wh_detail (по ID кластера)
            wm = get_warehouses_map()  # wid -> (wname, cid, cname)
            cid2name: Dict[int, str] = {}
            for _wid, (_wname, cid, cname) in wm.items():
                if int(cid or 0) and int(cid) not in cid2name:
                    cid2name[int(cid)] = (cname or f"Кластер {cid}").strip()

            if "ID кластера" in df_wh_detail.columns and "Кластер" not in df_wh_detail.columns:
                df_wh_detail["Кластер"] = df_wh_detail["ID кластера"].apply(
                    lambda c: cid2name.get(int(c) if pd.notna(c) else 0, f"Кластер {
                            int(c) if pd.notna(c) else 0}")
                )
                # чуть улучшим порядок колонок для восприятия
                preferred = ["Кластер", "Склад", "SKU", "Товар", "ID склада", "ID кластера"]
                rest = [c for c in df_wh_detail.columns if c not in preferred]
                df_wh_detail = df_wh_detail[preferred + rest]

            # ----- НОВАЯ: 3‑уровневая свёртка для листа «Склады»: Кластер → Склад → SKU
            def _with_outline_3level(
                df_src, top_col: str, mid_col: str, sum_cols: List[str]
            ) -> Tuple["pd.DataFrame", List[int]]:
                import pandas as pd  # type: ignore

                if top_col not in df_src.columns or mid_col not in df_src.columns:
                    return df_src, [0] * len(df_src)

                df = df_src.copy()

                # Сортировка: Кластер -> Склад -> (Действие asc, Рекомендация desc)
                order_action = {"Отгрузить": 0, "Поддерживать": 1, "Распродать": 2}
                if "Действие" in df.columns:
                    df["__act"] = df["Действие"].map(order_action).fillna(9)
                if "Рекомендация, шт" not in df.columns:
                    df["Рекомендация, шт"] = 0
                df.sort_values(
                    by=[top_col, mid_col, "__act", "Рекомендация, шт"],
                    ascending=[True, True, True, False],
                    inplace=True,
                )
                df.drop(columns=["__act"], inplace=True, errors="ignore")

                out_rows: List[dict] = []
                levels: List[int] = []
                cols = list(df.columns)

                def _sum_safe(frame, col):
                    try:
                        return float(pd.to_numeric(frame[col], errors="coerce").fillna(0).sum())
                    except Exception:
                        return None

                for top_val, top_df in df.groupby(top_col, sort=False):
                    # ИТОГО КЛАСТЕРА (уровень 0)
                    hdr_top = {c: "" for c in cols}
                    hdr_top[top_col] = top_val
                    if "Товар" in hdr_top:
                        hdr_top["Товар"] = f"{top_val} — итого кластера"
                    for c in sum_cols:
                        if c in top_df.columns:
                            hdr_top[c] = _sum_safe(top_df, c)
                    out_rows.append(hdr_top)
                    levels.append(0)

                    for mid_val, mid_df in top_df.groupby(mid_col, sort=False):
                        # ИТОГО СКЛАДА (уровень 1)
                        hdr_mid = {c: "" for c in cols}
                        hdr_mid[top_col] = top_val
                        hdr_mid[mid_col] = mid_val
                        if "Товар" in hdr_mid:
                            hdr_mid["Товар"] = f"{mid_val} — итого склада"
                        for c in sum_cols:
                            if c in mid_df.columns:
                                hdr_mid[c] = _sum_safe(mid_df, c)
                        out_rows.append(hdr_mid)
                        levels.append(1)

                        # Деталь SKU (уровень 2)
                        for _, r in mid_df.iterrows():
                            out_rows.append({c: r.get(c) for c in cols})
                            levels.append(2)

                df_out = pd.DataFrame(out_rows, columns=cols)
                return df_out, levels

            # Свёртка групп для листа «Кластеры» (как было)
            sum_cols = [
                "Рекомендация, шт",
                "На открытые, шт",
                "За закрытые, шт",
                "На лаг L, шт",
                "На лаг S, шт",
            ]
            df_cluster, lvl_cluster = _with_outline(df_cluster_detail, "Кластер", sum_cols)

            # Новая 3‑уровневая свёртка для листа «Склады»
            df_wh, lvl_wh = _with_outline_3level(df_wh_detail, "Кластер", "Склад", sum_cols)

            # Итог по листам
            df_cluster = _append_grand_total(df_cluster, sum_cols)
            df_wh = _append_grand_total(df_wh, sum_cols)

            # Запись листов
            df_sku.to_excel(wr, sheet_name="Товары", index=False)
            df_cluster.to_excel(wr, sheet_name="Кластеры", index=False)
            df_wh.to_excel(wr, sheet_name="Склады", index=False)

            # Форматирование
            if engine == "xlsxwriter":
                book = wr.book
                fmt_int = book.add_format({"num_format": "#,##0"})
                fmt_float = book.add_format({"num_format": "0.00"})
                fmt_header = book.add_format({"bold": True})
                fmt_total = book.add_format({"bold": True})
                fmt_group = book.add_format({"bold": True})

                def _format_sheet_xw(sh_name: str, df, levels: Optional[List[int]] = None):
                    sh = wr.sheets[sh_name]
                    cols = list(df.columns)
                    nrows, ncols = df.shape

                    def idx(name: str) -> Optional[int]:
                        try:
                            return cols.index(name)
                        except ValueError:
                            return None

                    def set_w(name: str, width: int, fmt=None):
                        i = idx(name)
                        if i is not None:
                            sh.set_column(i, i, width, fmt)

                    # Текстовые
                    set_w("SKU", 10)
                    set_w("Товар", 36)
                    set_w("Кластер", 28)
                    set_w("Склад", 32)
                    set_w("Действие", 16)
                    set_w("Маркер закрытого склада", 18)

                    # Числовые
                    for name in (
                        "ID кластера",
                        "ID склада",
                        "Срок доставки, дн",
                        "Срок до отгрузки, дн",
                        "База 30 дней, шт",
                        "План (30 дн), шт",
                        "Необходимо, шт",
                        "Запас, шт",
                        "Рекомендация, шт",
                        "На открытые, шт",
                        "За закрытые, шт",
                        "На лаг L, шт",
                        "На лаг S, шт",
                        "Закрытых складов, шт",
                    ):
                        set_w(name, 16, fmt_int)
                    set_w("Скорость, шт/день", 14, fmt_float)
                    set_w("Коэффициент", 12, fmt_float)
                    set_w("Обеспеченность, R", 12, fmt_float)

                    # Шапка жирным + автофильтр по всей таблице
                    sh.set_row(0, None, fmt_header)
                    if nrows >= 1 and ncols >= 1:
                        sh.autofilter(0, 0, nrows, ncols - 1)

                    # Заморозка: первая строка + две первые колонки (SKU, Товар)
                    sh.freeze_panes(1, 2)

                    # Свёртка групп (если передали уровни)
                    if levels:
                        # показать символы структуры
                        sh.outline_settings(True, True, True, False)

                        # Уровень 2 (SKU) — прячем
                        for i, lvl in enumerate(levels, start=1):  # +1 из-за заголовка
                            if lvl == 2:
                                sh.set_row(i, None, None, {"level": 2, "hidden": True})

                        # Уровень 1 (Склад) — прячем и отмечаем collapsed, чтобы у кластера был «+»
                        for i, lvl in enumerate(levels, start=1):
                            if lvl == 1:
                                sh.set_row(
                                    i,
                                    None,
                                    fmt_group,
                                    {"level": 1, "hidden": True, "collapsed": True},
                                )

                        # Уровень 0 (Кластер) — оставляем видимым, но collapsed=True для плюса
                        for i, lvl in enumerate(levels, start=1):
                            if lvl == 0:
                                sh.set_row(i, None, fmt_group, {"collapsed": True})

                    # Подсветить последнюю строку (ИТОГО ПО ЛИСТУ), если есть
                    if nrows >= 1:
                        sh.set_row(nrows, None, fmt_total)

                _format_sheet_xw("Товары", df_sku)
                _format_sheet_xw("Кластеры", df_cluster, lvl_cluster)
                _format_sheet_xw("Склады", df_wh, lvl_wh)

            else:
                # openpyxl
                try:
                    from openpyxl.utils import get_column_letter
                    from openpyxl.styles import Font

                    def _format_sheet_opx(sh_name: str, df, levels: Optional[List[int]] = None):
                        ws = wr.sheets[sh_name]
                        cols = list(df.columns)
                        nrows, ncols = df.shape
                        widths = {
                            "SKU": 10,
                            "Товар": 36,
                            "Кластер": 28,
                            "Склад": 32,
                            "ID кластера": 16,
                            "ID склада": 16,
                            "Скорость, шт/день": 14,
                            "Срок доставки, дн": 14,
                            "Срок до отгрузки, дн": 14,
                            "План (30 дн), шт": 16,
                            "Коэффициент": 12,
                            "База 30 дней, шт": 16,
                            "Необходимо, шт": 16,
                            "Запас, шт": 16,
                            "Обеспеченность, R": 12,
                            "Действие": 16,
                            "Рекомендация, шт": 16,
                            "На открытые, шт": 16,
                            "За закрытые, шт": 16,
                            "На лаг L, шт": 16,
                            "На лаг S, шт": 16,
                            "Закрытых складов, шт": 16,
                            "Маркер закрытого склада": 18,
                        }
                        for i, name in enumerate(cols, start=1):
                            w = widths.get(name, 16 if i > 2 else 12)
                            ws.column_dimensions[get_column_letter(i)].width = w

                        # freeze panes: C2
                        ws.freeze_panes = "C2"

                        # Автофильтр по всем данным
                        if nrows >= 1 and ncols >= 1:
                            ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows + 1}"

                        # Жирная шапка
                        for cell in ws[1]:
                            cell.font = Font(bold=True)

                        # Свёртка групп (уровни 0/1/2), старт — свёрнуто (склады и SKU скрыты)
                        if levels:
                            ws.sheet_properties.outlinePr.summaryBelow = True
                            for i, lvl in enumerate(levels, start=2):  # данные со 2-й строки
                                rd = ws.row_dimensions[i]
                                if lvl == 2:
                                    rd.outlineLevel = 2
                                    rd.hidden = True
                                elif lvl == 1:
                                    rd.outlineLevel = 1
                                    rd.hidden = True
                                else:  # lvl == 0 (кластер)
                                    # заголовок кластера — видимый и жирный
                                    for c in ws[i]:
                                        c.font = Font(bold=True)

                        # Подсветить последнюю строку (ИТОГО ПО ЛИСТУ)
                        try:
                            for cell in ws[nrows + 1]:
                                cell.font = Font(bold=True)
                        except Exception:
                            pass

                    _format_sheet_opx("Товары", df_sku)
                    _format_sheet_opx("Кластеры", df_cluster, lvl_cluster)
                    _format_sheet_opx("Склады", df_wh, lvl_wh)
                except Exception:
                    pass

    last_err: Optional[Exception] = None
    for engine in ("xlsxwriter", "openpyxl"):
        try:
            _write_excel(engine)
            return path
        except Exception as e:
            last_err = e
            continue

    raise RuntimeError(
        f"Excel-экспорт не удался (xlsxwriter/openpyxl). Последняя ошибка: {last_err}"
    )


__all__ = [
    "compute_need",
    "format_need_text",
    "export_need_excel",
    "need_to_ship_text",
    "shipments_text",
    "need_text",
]
